import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.cell.cell import MergedCell
import zipfile
import re


def try_open_excel(path):
    candidates = [path]
    if not os.path.splitext(path)[1]:
        candidates += [path + ext for ext in (".xlsx", ".xls", ".xlsm")]
    for p in candidates:
        if os.path.exists(p):
            try:
                if p.endswith(('.xlsx', '.xlsm')):
                    with zipfile.ZipFile(p, 'r') as zf:
                        pass
                return p
            except (zipfile.BadZipFile, Exception):
                continue
    return None


def is_valid_excel_file(file_path):
    """Check if the file is a valid Excel file"""
    try:
        if file_path.endswith(('.xlsx', '.xlsm')):
            with zipfile.ZipFile(file_path, 'r') as zf:
                return True
        elif file_path.endswith('.xls'):
            pd.read_excel(file_path, nrows=1, engine='xlrd')
            return True
        else:
            pd.read_excel(file_path, nrows=1)
            return True
    except Exception as e:
        print(f"❌ File validation failed for {file_path}: {e}")
        return False


def get_cell_type(cellname):
    """Determine cell type based on cellname prefix"""
    if pd.isna(cellname):
        return "unknown"

    cellname_str = str(cellname).strip().upper()

    if cellname_str.startswith('Y'):
        return "FDD"
    elif cellname_str.startswith('Q'):
        return "TDD"
    else:
        return "unknown"


def get_operator_from_cellname(cellname):
    """Determine operator (BYT/SFR) based on last letter of cell name"""
    if pd.isna(cellname):
        return "unknown"
    
    cellname_str = str(cellname).strip().upper()
    
    if len(cellname_str) == 0:
        return "unknown"
    
    last_letter = cellname_str[-1]
    
    # BYT cells end with A/B/C/D/E/F
    if last_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        return "BYT"
    # SFR cells end with N/O/P/Q/R/S
    elif last_letter in ['N', 'O', 'P', 'Q', 'R', 'S']:
        return "SFR"
    else:
        return "unknown"


def load_nename_categories(category_file_path):
    """Load NeName categories from Excel file"""
    try:
        print(f"📖 Loading NeName categories from: {category_file_path}")
        category_df = pd.read_excel(category_file_path, engine='openpyxl')
        
        # Create a mapping dictionary for NeName to categories
        print(f"Available columns categories : {list(category_df.columns)}")
        nename_categories = {}
        
        required_columns = ['NeName', 'Type', 'Cell', 'Gen', 'Remarque']
        missing_columns = [col for col in required_columns if col not in category_df.columns]
        
        if missing_columns:
            print(f"❌ Missing columns in category file: {missing_columns}")
            print(f"Available columns categories : {list(category_df.columns)}")
            return {}
        
        for _, row in category_df.iterrows():
            if pd.isna(row['NeName']):
                continue
                
            nename = str(row['NeName']).strip()
            nename_categories[nename] = {
                'Type': row['Type'] if pd.notna(row['Type']) else '',
                'Cell': row['Cell'] if pd.notna(row['Cell']) else '',
                'Gen': row['Gen'] if pd.notna(row['Gen']) else '',
                'Remarque': row['Remarque'] if pd.notna(row['Remarque']) else ''
            }
        
        print(f"✅ Loaded categories for {len(nename_categories)} NeNames")
        
            
        return nename_categories
        
    except Exception as e:
        print(f"❌ Error loading NeName categories: {e}")
        import traceback
        traceback.print_exc()
        return {}


def get_node_type(nename, nename_categories):
    """Determine node type based on NeName using the categories file"""
    if pd.isna(nename):
        return "unknown"
    
    nename_str = str(nename).strip()
    
    # Check if we have this NeName in our categories
    if nename_str in nename_categories:
        category_info = nename_categories[nename_str]
        cell_config = category_info.get('Cell', '')
        netype = category_info.get('Type', '')
        
        # First check for TDD+FDD co-nodes
        if cell_config == "TDD + FDD" or cell_config == "TDD+FDD":
            return "TDD+FDD"
        
        # Then use the Type column for categorization
        if netype:
            return netype
    
    # Fallback to original logic for unknown NeNames
    return "unknown"


def is_gen2_cell(gen_value):
    """Check if cell is Gen2 based on Gen column"""
    if pd.isna(gen_value):
        return False
    gen_str = str(gen_value).strip().upper()
    return gen_str == "GEN2" or gen_str == "2" or "GEN2" in gen_str


def is_aas_cell(cellname):
    """Check if cell is an AAS cell (only Q90265A specifically)"""
    if pd.isna(cellname):
        return False
    
    cellname_str = str(cellname).strip().upper()
    
    # Only Q90265A is considered AAS, not all Q...A cells
    aas_cells = [
        "Q90265A",
        # Add more specific AAS cell names here if needed
    ]
    
    return cellname_str in aas_cells


def normalize_actual_value(actual_value):
    """Convert only FAUX/VRAI to false/true, handle Python booleans, remove .0 from integers"""
    if pd.isna(actual_value) or actual_value == "":
        return None
    
    # Handle Python boolean objects - convert to lowercase
    if isinstance(actual_value, bool):
        return "true" if actual_value else "false"
    
    value_str = str(actual_value).strip()
    
    # Convert only French boolean values and Python-style booleans
    if value_str.upper() in ['VRAI', 'TRUE']:
        return "true"
    elif value_str.upper() in ['FAUX', 'FALSE']:
        return "false"
    
    # Remove .0 from integer values (e.g., 8.0 -> 8)
    try:
        # Check if it's a float that represents an integer
        if '.' in value_str:
            float_val = float(value_str)
            if float_val.is_integer():
                return str(int(float_val))
    except (ValueError, TypeError):
        pass
    
    # Keep all other values as they are
    return value_str


def convert_for_display(value):
    """Convert for display - ensure true/false are lowercase"""
    if pd.isna(value) or value == "":
        return value
    
    # Handle Python boolean objects
    if isinstance(value, bool):
        return "true" if value else "false"
    
    value_str = str(value).strip()
    
    # Convert boolean values to lowercase for display
    if value_str.upper() in ['VRAI', 'TRUE']:
        return "true"
    elif value_str.upper() in ['FAUX', 'FALSE']:
        return "false"
    
    # Keep all other values as they are
    return value_str


def is_na_value(value):
    """Check if value is N/A, null, empty, or similar"""
    if pd.isna(value) or value == "":
        return True

    value_str = str(value).strip().lower()
    na_values = ['n/a', 'null', 'none', 'nan', 'empty', 'vide', '-', 'read-only']
    return value_str in na_values


def extract_main_value(value):
    """Extract main value before space or ="""
    if pd.isna(value) or value == "" or is_na_value(value):
        return None

    value_str = str(value).strip()

    # If value contains space or =, extract the first part
    if ' ' in value_str or '=' in value_str:
        # Split by space or = and take the first part
        parts = re.split(r'[\s=]', value_str)
        for part in parts:
            part = part.strip()
            if part and part != '':  # Skip empty parts
                return part

    return value_str


def normalize_parameter_key(key):
    """Normalize parameter key by removing common prefixes and converting to lowercase"""
    if pd.isna(key):
        return ""

    key_str = str(key).strip().lower()

    # Remove common telecom parameter prefixes
    prefixes = [
        'vsdata', 'vs', 'data', 'nr', 'lte', 'cell', 'param', 'parameter',
        'eutran', 'geran', 'uran', 'wcdma', 'gsm', 'umts', 'hspa'
    ]

    # Remove prefixes
    for prefix in prefixes:
        if key_str.startswith(prefix):
            key_str = key_str[len(prefix):]
            break

    return key_str.strip()


def fuzzy_key_match(expected_key, actual_key):
    """Check if keys match approximately (handles prefixes like vsData)"""
    if pd.isna(expected_key) or pd.isna(actual_key):
        return False

    expected_clean = normalize_parameter_key(expected_key)
    actual_clean = normalize_parameter_key(actual_key)

    # Exact match after normalization
    if expected_clean == actual_clean:
        return True

    # Check if one is contained in the other (for partial matches)
    if expected_clean in actual_clean or actual_clean in expected_clean:
        return True

    # Split by common separators and check word overlap
    expected_words = set(re.findall(r'[a-z]+', expected_clean))
    actual_words = set(re.findall(r'[a-z]+', actual_clean))

    if expected_words and actual_words:
        common_words = expected_words & actual_words
        # If most words match, consider it a match
        if len(common_words) >= min(len(expected_words), len(actual_words)):
            return True

    return False


def parse_key_value_pairs(value_str):
    """Parse key=value pairs from a string with robust error handling"""
    if pd.isna(value_str) or value_str == "":
        return {}

    pairs = {}
    try:
        # Split by comma and parse each pair
        for pair in str(value_str).split(','):
            pair = pair.strip()
            if '=' in pair:
                key, value = pair.split('=', 1)
                pairs[key.strip()] = value.strip()
    except Exception as e:
        print(f"⚠️  Error parsing key-value pairs: {e}")

    return pairs


def find_key_value_in_string(search_key, search_value, long_string):
    """Search for a specific key=value pair within a longer string containing multiple pairs"""
    if pd.isna(long_string) or long_string == "":
        return False

    long_str = str(long_string).strip()

    # Handle the case where the expected value is a simple key=value pair
    # and the actual value is a longer path containing that key=value pair
    if "=" in long_str and "," in long_str:
        # This looks like a comma-separated list of key=value pairs
        all_pairs = parse_key_value_pairs(long_str)

        # Search for a matching key (with fuzzy matching) and exact value match
        for actual_key, actual_value in all_pairs.items():
            if fuzzy_key_match(search_key, actual_key) and actual_value == search_value:
                return True
    else:
        # Handle case where the actual value might be a single key=value pair or path
        # Check if the search_key=search_value appears anywhere in the long string
        expected_pair = f"{search_key}={search_value}"
        if expected_pair in long_str:
            return True

        # Also try with fuzzy key matching
        # Split the long string by commas and check each part
        parts = [part.strip() for part in long_str.split(',')]
        for part in parts:
            if '=' in part:
                actual_key, actual_value = part.split('=', 1)
                actual_key = actual_key.strip()
                actual_value = actual_value.strip()
                if fuzzy_key_match(search_key, actual_key) and actual_value == search_value:
                    return True

    return False


def validate_cell_local_id(actual_value, cellname):
    """Custom validation for cellLocalId parameter based on cell name pattern"""
    if pd.isna(actual_value) or pd.isna(cellname):
        return False
    
    actual_str = str(actual_value).strip()
    cellname_str = str(cellname).strip().upper()
    
    
    # Define the mapping based on cell name patterns
    cell_mapping = {
        # TDD cells (starting with Q)
        'Q': {
            'A': '51', 'B': '52', 'C': '53',
            'D': '151', 'E': '152', 'F': '153', 
            'N': '72', 'O': '73', 'P': '74',
            'Q': '172', 'R': '173', 'S': '174'
        },
        # FDD cells (starting with Y)  
        'Y': {
            'A': '66', 'B': '67', 'C': '68',
            'D': '166', 'E': '167', 'F': '168',
            'N': '69', 'O': '70', 'P': '71',
            'Q': '169', 'R': '170', 'S': '171'
        }
    }
    
    # Extract the last character (sector identifier)
    if len(cellname_str) >= 1:
        cell_prefix = cellname_str[0]  # Q or Y
        sector_id = cellname_str[-1]   # A, B, C, etc.
        
        if cell_prefix in cell_mapping and sector_id in cell_mapping[cell_prefix]:
            expected_value = cell_mapping[cell_prefix][sector_id]
            return actual_str == expected_value
    
    print(f"   ❌ No mapping found for cellname: '{cellname_str}'")
    return False


def get_expected_cell_local_id(cellname):
    """Get the expected cellLocalId value for display purposes"""
    if pd.isna(cellname):
        return "Unknown"
    
    cellname_str = str(cellname).strip().upper()
    
    # Define the mapping based on cell name patterns
    cell_mapping = {
        'Q': {
            'A': '51', 'B': '52', 'C': '53',
            'D': '151', 'E': '152', 'F': '153', 
            'N': '72', 'O': '73', 'P': '74',
            'Q': '172', 'R': '173', 'S': '174'
        },
        'Y': {
            'A': '66', 'B': '67', 'C': '68',
            'D': '166', 'E': '167', 'F': '168',
            'N': '69', 'O': '70', 'P': '71',
            'Q': '169', 'R': '170', 'S': '171'
        }
    }
    
    # Extract the last character (sector identifier)
    if len(cellname_str) >= 1:
        cell_prefix = cellname_str[0]  # Q or Y
        sector_id = cellname_str[-1]   # A, B, C, etc.
        
        if cell_prefix in cell_mapping and sector_id in cell_mapping[cell_prefix]:
            return cell_mapping[cell_prefix][sector_id]
    
    return "Unknown mapping"


def validate_operator_specific_value(actual_value, expected_value, operateur, cell_type):
    """Validate operator-specific values (BYT vs SFR) - ENHANCED"""
    if pd.isna(actual_value) or pd.isna(expected_value):
        return False
        
    actual_normalized = normalize_actual_value(actual_value)
    expected_str = str(expected_value).strip()
    
    
    # Handle patterns like "BYT = 640320 SFR = 635328" and "BYT = 426990 SFR = 422930"
    if "BYT" in expected_str.upper() and "SFR" in expected_str.upper():
        # Extract BYT and SFR values using more flexible patterns
        byt_match = re.search(r'BYT\s*[=:]\s*(\d+)', expected_str, re.IGNORECASE)
        sfr_match = re.search(r'SFR\s*[=:]\s*(\d+)', expected_str, re.IGNORECASE)
        
        if byt_match and sfr_match:
            byt_value = byt_match.group(1)
            sfr_value = sfr_match.group(1)
            
            # Check based on operator and cell type
            if operateur == "BYT":
                expected_value = byt_value
            elif operateur == "SFR":
                expected_value = sfr_value
            else:
                # If operator unknown, use BYT as default for TDD, SFR for FDD
                if cell_type == "TDD":
                    expected_value = byt_value
                else:
                    expected_value = sfr_value
            
            return actual_normalized == expected_value
    
    # Handle operator-specific patterns like "BWPSet=1 (BYT) / =11 (SFR)"
    if "(BYT)" in expected_str and "(SFR)" in expected_str:
        # Extract BYT and SFR values
        byt_match = re.search(r'=(\d+)\s*\(BYT\)', expected_str)
        sfr_match = re.search(r'=(\d+)\s*\(SFR\)', expected_str)
        
        if byt_match and sfr_match:
            byt_value = byt_match.group(1)
            sfr_value = sfr_match.group(1)
            
            # Check based on operator
            if operateur == "BYT":
                expected_value = byt_value
            elif operateur == "SFR":
                expected_value = sfr_value
            else:
                # If operator unknown, use BYT as default
                expected_value = byt_value
            
            # For operator-specific values, check if the expected value is contained in the actual value
            # This handles cases like actual="SubNetwork=...,vsDataBWPSet=1" and expected="1"
            if expected_value in actual_normalized:
                return True
            else:
                return False
    
    return False


def validate_type_specific_value(actual_value, expected_value, remarque, operateur, cell_type, param_name):
    """Validate type-specific values (ZTD, CRZ, etc.) and operator-specific values - FIXED"""
    if pd.isna(actual_value) or pd.isna(expected_value):
        return False
        
    actual_str = str(actual_value).strip()
    expected_str = str(expected_value).strip()
    
    
    # Handle multiple type-specific values like "0 = 0dB en ZTD 3 = 3dB en CRZ"
    if "en ZTD" in expected_str and "en CRZ" in expected_str:
        # Extract values for both ZTD and CRZ
        ztd_match = re.search(r'(\d+)\s*=\s*\d+dB\s*en ZTD', expected_str)
        crz_match = re.search(r'(\d+)\s*=\s*\d+dB\s*en CRZ', expected_str)
        
        if ztd_match and crz_match:
            ztd_value = ztd_match.group(1)
            crz_value = crz_match.group(1)
            
            if "ZTD" in str(remarque) and cell_type == "TDD":
                print(f"   ZTD TDD site - expecting: '{ztd_value}'")
                return actual_str == ztd_value
            elif "CRZ" in str(remarque) and cell_type == "TDD":
                print(f"   CRZ TDD site - expecting: '{crz_value}'")
                return actual_str == crz_value
            else:
                # For other types, use default value
                return False
    
    # Special handling for raResponseWindow with "en ZTD" pattern
    if param_name and "raresponsewindow" in param_name.lower() and "en ZTD" in expected_str:
        if "ZTD" in str(remarque) and cell_type == "TDD":
            # For ZTD TDD sites, expect the main value (20)
            main_value = extract_main_value(expected_value)
            print(f"   ZTD TDD site - expecting: '{main_value}'")
            return actual_str == main_value
        else:
            # For non-ZTD sites or non-TDD cells, use default value
            return False  # Will be handled by default value logic
    
    # Handle type specifications like "0 = 0dB en ZTD" (single type)
    if "en ZTD" in expected_str:
        if "ZTD" in str(remarque):
            # For ZTD sites, expect the main value
            main_value = extract_main_value(expected_value)
            print(f"   ZTD site - expecting: '{main_value}'")
            return actual_str == main_value
        else:
            # For non-ZTD sites (like CRZ), check if there's a CRZ-specific value
            if "CRZ" in str(remarque):
                # Look for CRZ-specific value in the pattern
                crz_match = re.search(r'(\d+)\s*=\s*\d+dB\s*en CRZ', expected_str)
                if crz_match:
                    crz_value = crz_match.group(1)
                    print(f"   CRZ site - expecting: '{crz_value}'")
                    return actual_str == crz_value
            # Default for non-ZTD when no specific value mentioned
            return False
    
    # Handle type specifications like "3 = 3dB en CRZ" (single type)
    if "en CRZ" in expected_str:
        if "CRZ" in str(remarque):
            # For CRZ sites, expect the main value
            main_value = extract_main_value(expected_value)
            print(f"   CRZ site - expecting: '{main_value}'")
            return actual_str == main_value
        else:
            # For non-CRZ sites, check if there's a ZTD-specific value
            if "ZTD" in str(remarque):
                # Look for ZTD-specific value in the pattern
                ztd_match = re.search(r'(\d+)\s*=\s*\d+dB\s*en ZTD', expected_str)
                if ztd_match:
                    ztd_value = ztd_match.group(1)
                    print(f"   ZTD site - expecting: '{ztd_value}'")
                    return actual_str == ztd_value
            # Default for non-CRZ when no specific value mentioned
            return False
    
    # Handle other type specifications
    type_mappings = {
        "en ZTD": "ZTD",
        "en CRZ": "CRZ", 
        "ZTD": "ZTD",
        "CRZ": "CRZ",
        "Ran4": "Ran4"
    }
    
    for pattern, type_name in type_mappings.items():
        if pattern in expected_str:
            if type_name in str(remarque):
                main_value = extract_main_value(expected_value)
                print(f"   {type_name} site - expecting: '{main_value}'")
                return actual_str == main_value
            else:
                # For non-matching types, use main value as default
                main_value = extract_main_value(expected_value)
                print(f"   Non-{type_name} site - expecting: '{main_value}'")
                return actual_str == main_value
    
    return False


def detect_validation_pattern(expected_value, param_name=""):
    """Detect what type of validation pattern the expected value represents - ENHANCED"""
    if pd.isna(expected_value) or is_na_value(expected_value):
        return "no_expected_value"

    expected_str = str(expected_value).strip()

    # Operator-specific pattern with BYT/SFR values
    if ("BYT" in expected_str.upper() and "SFR" in expected_str.upper() and 
        (re.search(r'BYT\s*[=:]\s*\d+', expected_str, re.IGNORECASE) and 
         re.search(r'SFR\s*[=:]\s*\d+', expected_str, re.IGNORECASE))):
        return "operator_specific"
    
    # Operator-specific pattern (BYT/SFR)
    if "(BYT)" in expected_str and "(SFR)" in expected_str:
        return "operator_specific"
    
    # Multiple options with operator info
    if "/" in expected_str and ("BYT" in expected_str.upper() or "SFR" in expected_str.upper()):
        return "operator_specific"

    # Type-specific pattern (ZTD, CRZ, etc.) - especially for multiple types like "0 = 0dB en ZTD 3 = 3dB en CRZ"
    type_indicators = ["en ZTD", "en CRZ", "ZTD", "CRZ", "Ran4", "en SFR"]
    if any(indicator in expected_str for indicator in type_indicators):
        # Check if there are multiple type specifications
        if ("en ZTD" in expected_str and "en CRZ" in expected_str) or \
           (expected_str.count("en ZTD") > 1) or \
           (expected_str.count("en CRZ") > 1):
            return "node_specific_multiple"
        else:
            return "node_specific"

    # Complex key-value pairs (like EnergyEfficiency=1,EnergyOptPwrAlloc=Default)
    if "," in expected_str and "=" in expected_str:
        pairs = [pair.strip() for pair in expected_str.split(",")]
        if all("=" in pair for pair in pairs):
            return "key_value_pairs"

    # Value with explanation (e.g., "0 = NO_LOCK", "20 = 20 slots")
    if (" = " in expected_str or " " in expected_str) and len(expected_str.split()) >= 2:
        first_part = expected_str.split()[0]
        number_pattern = r'^-?\d+\.?\d*$'
        if re.match(number_pattern, first_part) or first_part.lower() in ['true', 'false']:
            return "value_with_explanation"

    # Single key-value pair
    if "=" in expected_str and "," not in expected_str:
        return "single_key_value"

    partial_keywords = ['enabled', 'disabled', 'active', 'inactive', 'on', 'off', 'yes', 'no']
    if any(keyword in expected_str.lower() for keyword in partial_keywords):
        return "partial_match"

    return "exact_match"


def apply_special_validation(expected_value, actual_value, pattern_type, node_type=None, 
                           expected_co_node_value=None, cell_type=None, operateur=None, remarque=None, param_name=None):
    """Apply special validation based on the detected pattern - FIXED"""
    if pd.isna(actual_value) or pd.isna(expected_value) or is_na_value(expected_value):
        return False

    expected_str = str(expected_value).strip()
    actual_normalized = normalize_actual_value(actual_value)
    expected_normalized = normalize_actual_value(expected_value)

    if pattern_type == "operator_specific":
        return validate_operator_specific_value(actual_value, expected_value, operateur, cell_type)

    elif pattern_type == "node_specific_multiple":
        # Handle multiple type-specific values like "0 = 0dB en ZTD 3 = 3dB en CRZ"
        return validate_type_specific_value(actual_value, expected_value, remarque, operateur, cell_type, param_name)

    elif pattern_type == "node_specific":
        return validate_type_specific_value(actual_value, expected_value, remarque, operateur, cell_type, param_name)

    elif pattern_type == "value_with_explanation":
        main_value = extract_main_value(expected_str)
        main_value_normalized = normalize_actual_value(main_value)
        return actual_normalized == main_value_normalized

    elif pattern_type == "key_value_pairs":
        expected_pairs = parse_key_value_pairs(expected_str)
        for exp_key, exp_value in expected_pairs.items():
            exp_value_normalized = normalize_actual_value(exp_value)
            if not find_key_value_in_string(exp_key, exp_value_normalized, actual_normalized):
                return False
        return True

    elif pattern_type == "single_key_value":
        if "=" in expected_str:
            exp_key, exp_value = expected_str.split("=", 1)
            exp_key = exp_key.strip()
            exp_value = exp_value.strip()
            exp_value_normalized = normalize_actual_value(exp_value)
            result = find_key_value_in_string(exp_key, exp_value_normalized, actual_normalized)
            return result
        return False

    elif pattern_type == "partial_match":
        result = expected_str.lower() in str(actual_normalized).lower()
        print(f"   Partial match: '{expected_str}' in '{actual_normalized}' -> {result}")
        return result

    return False

def get_display_expected_value(param_info, cell_type, node_type, expected_co_node_value, param_name, cellname, operateur):
    """Get the expected value for display purposes"""
    # Special handling for cellLocalId
    if param_name and "celllocalid" in param_name.lower():
        return get_expected_cell_local_id(cellname)
    
    if node_type == "TDD+FDD" and not pd.isna(expected_co_node_value) and not is_na_value(expected_co_node_value):
        return expected_co_node_value
    elif cell_type == "FDD":
        if not pd.isna(param_info["Valeur Bytel FDD ESS 15MHz"]) and not is_na_value(param_info["Valeur Bytel FDD ESS 15MHz"]):
            return param_info["Valeur Bytel FDD ESS 15MHz"]
        elif not pd.isna(param_info["Valeur par défaut RBS"]) and not is_na_value(param_info["Valeur par défaut RBS"]):
            return param_info["Valeur par défaut RBS"]
        else:
            return param_info["Valeur Bytel TDD MidBand"] or param_info["Valeur Bytel TDD HigBand"]
    else:  # TDD or unknown
        if not pd.isna(param_info["Valeur Bytel TDD MidBand"]) and not is_na_value(param_info["Valeur Bytel TDD MidBand"]):
            return param_info["Valeur Bytel TDD MidBand"]
        elif not pd.isna(param_info["Valeur Bytel TDD HigBand"]) and not is_na_value(param_info["Valeur Bytel TDD HigBand"]):
            return param_info["Valeur Bytel TDD HigBand"]
        elif not pd.isna(param_info["Valeur par défaut RBS"]) and not is_na_value(param_info["Valeur par défaut RBS"]):
            return param_info["Valeur par défaut RBS"]
        else:
            return "No expected value"


# Add this list of parameters to exclude from validation
EXCLUDED_PARAMETERS = [
    "endcDlLegSwitchEnabled",
    "endcDlNrLowQualThresh", 
    "endcDlNrQualHyst",
    "endcUlLegSwitchEnabled",
    "endcUlNrLowQualThresh",
    "endcUlNrQualHyst",
    "fddLteCoexistence",
    "dlRobustLaEnabled",
    "drxProfileEnabled",
    "trafficGroupRef",
    "ulRobustLaEnabled",
    "nCGI",
    "nRPCI"
]

def validate_parameter_value(actual_value, expected_tdd_value, expected_fdd_value, expected_default_value, expected_co_node_value, cell_type,
                             parameter_name, node_type, operateur, remarque, cellname, gen_value):
    """Validate if the actual value matches the expected value - ENHANCED"""

    # Skip validation for administrativeState parameter
    if parameter_name and "administrativestate" in parameter_name.lower():
        return "skipped"

    if parameter_name and "nrtac" in parameter_name.lower():
        return "skipped"

    # Skip excluded parameters
    if parameter_name and parameter_name in EXCLUDED_PARAMETERS:
        return "skipped"

    # Special handling for Gen2 cells - if all Gen2 cells have empty values, mark as correct
    if is_gen2_cell(gen_value) and (pd.isna(actual_value) or actual_value == ""):
        return "correct"

    # Special handling for cellLocalId parameter
    if parameter_name and "celllocalid" in parameter_name.lower():
        if pd.isna(actual_value) or actual_value == "":
            return "no_data"
        if validate_cell_local_id(actual_value, cellname):
            return "correct"
        else:
            return "incorrect"

    # Special handling for rim parameters - only use TDD value for AAS cells
    if parameter_name and "rim" in parameter_name.lower() and cell_type == "TDD":
        if not is_aas_cell(cellname):
            # For non-AAS TDD cells, use default value instead of TDD value
            expected_tdd_value = None

    # Determine which expected value to use based on cell type
    expected_value = None
    
    # For TDD cells
    if cell_type == "TDD":
        if not pd.isna(expected_tdd_value) and not is_na_value(expected_tdd_value):
            expected_value = expected_tdd_value
        elif not pd.isna(expected_default_value) and not is_na_value(expected_default_value):
            expected_value = expected_default_value
        else:
            # All expected values are empty/null - expect actual value to also be empty
            expected_value = ""
    
    # For FDD cells  
    elif cell_type == "FDD":
        if not pd.isna(expected_fdd_value) and not is_na_value(expected_fdd_value):
            expected_value = expected_fdd_value
        elif not pd.isna(expected_default_value) and not is_na_value(expected_default_value):
            expected_value = expected_default_value
        else:
            # All expected values are empty/null - expect actual value to also be empty
            expected_value = ""
    
    # For unknown cell types, use default if available
    else:
        if not pd.isna(expected_default_value) and not is_na_value(expected_default_value):
            expected_value = expected_default_value
        else:
            # All expected values are empty/null - expect actual value to also be empty
            expected_value = ""

    # For TDD+FDD co-nodes, prioritize the co-node value
    if node_type == "TDD+FDD" and not pd.isna(expected_co_node_value) and not is_na_value(expected_co_node_value):
        expected_value = expected_co_node_value

    # Handle case where all expected values are empty
    if expected_value == "":
        if pd.isna(actual_value) or actual_value == "":
            return "correct"  # Both expected and actual are empty - correct!
        else:
            return "incorrect"  # Expected empty but actual has value - incorrect!

    # If no valid expected value found and it's not empty case
    if expected_value is None:
        return "no_expected_value"

    # If actual value is empty but expected has value
    if pd.isna(actual_value) or actual_value == "":
        return "incorrect"  # Expected value but actual is empty - incorrect!

    # Normalize values
    actual_normalized = normalize_actual_value(actual_value)
    expected_str = str(expected_value).strip()

    # First, try special validation patterns
    pattern_type = detect_validation_pattern(expected_value, parameter_name)
    
    # Handle operator-specific patterns
    if pattern_type == "operator_specific":
        if validate_operator_specific_value(actual_value, expected_value, operateur, cell_type):
            return "correct_fuzzy"
    
    # Handle type-specific patterns (ZTD, CRZ, etc.)
    if pattern_type == "node_specific":
        if validate_type_specific_value(actual_value, expected_value, remarque, operateur, cell_type, parameter_name):
            return "correct_fuzzy"

    # Try special validation
    if pattern_type != "exact_match" and pattern_type != "no_expected_value":
        if apply_special_validation(expected_value, actual_value, pattern_type, node_type, expected_co_node_value, cell_type, operateur, remarque, parameter_name):
            return "correct_fuzzy"

    # Method 1: Compare with extracted main value
    expected_main = extract_main_value(expected_value)
    if expected_main:
        expected_main_normalized = normalize_actual_value(expected_main)
        if actual_normalized == expected_main_normalized:
            return "correct_extracted"

    # Method 2: Direct comparison
    expected_normalized = normalize_actual_value(expected_value)
    if actual_normalized == expected_normalized:
        print(f"   ✅ Correct (exact match)")
        return "correct"

    # Method 3: For complex values like "EnergyEfficiency=1,EnergyOptPwrAlloc=Default"
    if "," in expected_str and "=" in expected_str:
        expected_pairs = parse_key_value_pairs(expected_str)
        all_pairs_found = True
        for exp_key, exp_value in expected_pairs.items():
            if not find_key_value_in_string(exp_key, exp_value, actual_normalized):
                all_pairs_found = False
                break
        if all_pairs_found:
            print(f"   ✅ Correct (all key-value pairs found)")
            return "correct_fuzzy"

    # Method 4: Try numeric comparison
    try:
        actual_num = float(actual_normalized)
        expected_num = float(expected_normalized)
        if actual_num == expected_num:
            print(f"   ✅ Correct (numeric match)")
            return "correct_numeric"
    except (ValueError, TypeError):
        pass

    return "incorrect"


# Get user choice
choice = input("Select sheet type (NRCellCU / NRCellDU): ").strip()
if choice not in ["NRCellCU", "NRCellDU"]:
    print("❌ Invalid choice. Please choose NRCellCU or NRCellDU.")
    raise SystemExit

SHEET_NAME = choice

# Get NeName category file path
print("\n📁 Please provide the NeName category Excel file (Type_Site_ENM.xlsx)")
category_file = input("Enter path to NeName category Excel file: ").strip()
category_path = try_open_excel(category_file)

if category_path is None:
    print(f"❌ NeName category file not found: {category_file}")
    raise SystemExit

# Load NeName categories
nename_categories = load_nename_categories(category_path)
if not nename_categories:
    print("❌ Failed to load NeName categories. Please check the file format.")
    raise SystemExit

# Get parameter file path
print("\n📁 Please provide the parameter Excel file")
param_file = input("Enter path to parameter Excel file (updated file): ").strip()
param_path = try_open_excel(param_file)

if param_path is None:
    print(f"❌ Parameter file not found: {param_file}")
    raise SystemExit

# Get data file path
print("\n📁 Please provide the data Excel file")
data_file = input("Enter path to data Excel file (with node data): ").strip()
data_path = try_open_excel(data_file)

if data_path is None:
    print(f"❌ Data file not found: {data_file}")
    raise SystemExit

print("🔍 Validating Excel files...")
if not is_valid_excel_file(param_path):
    print(f"❌ Parameter file is not a valid Excel file: {param_path}")
    raise SystemExit

if not is_valid_excel_file(data_path):
    print(f"❌ Data file is not a valid Excel file: {data_path}")
    raise SystemExit

if not is_valid_excel_file(category_path):
    print(f"❌ NeName category file is not a valid Excel file: {category_path}")
    raise SystemExit

print(f"📁 Parameter file: {param_path}")
print(f"📁 Data file: {data_path}")
print(f"📁 NeName category file: {category_path}")

# Load parameter workbook
try:
    print("📖 Loading parameter workbook...")
    param_wb = load_workbook(param_path)
except Exception as e:
    print(f"❌ Error loading parameter workbook: {e}")
    raise SystemExit

# Check if selected sheet exists
if SHEET_NAME not in param_wb.sheetnames:
    print(f"❌ Sheet '{SHEET_NAME}' not found in parameter file")
    print(f"Available sheets: {param_wb.sheetnames}")
    raise SystemExit

# Read parameter sheet
param_ws = param_wb[SHEET_NAME]
param_header = [cell.value for cell in param_ws[1]]

print(f"📋 Columns in parameter sheet: {param_header}")

# Map column names to indices
param_col_idx_map = {name: idx + 1 for idx, name in enumerate(param_header) if name is not None}

# Check required columns in parameter file
required_param_columns = ["Parameter", "Valeur par défaut RBS", "Valeur Bytel TDD MidBand",
                          "Valeur Bytel FDD ESS 15MHz", "Valeur Bytel TDD HigBand"]

# Add TDD+FDD co-node column for NRCellCU
if SHEET_NAME == "NRCellCU":
    required_param_columns.append("Valeur Bytel TDD+FDD co-node\nAppliquer la valeur commune si valeur TDD et FDD sont même, sinon appliquer la valeur spécifiée dans cette colonne.")

missing_columns = []
for col in required_param_columns:
    if col not in param_col_idx_map:
        missing_columns.append(col)

if missing_columns:
    print(f"❌ Missing required columns in parameter file: {missing_columns}")
    raise SystemExit

# Get parameters (excluding read-only ones and administrativeState)
parameters_to_include = []
parameter_data = {}  # Store parameter info
missing_parameters_in_data = []  # Track parameters not found in data file

param_col_idx = param_col_idx_map["Parameter"]
readonly_col_idx = param_col_idx_map["Valeur par défaut RBS"]

print(f"\n🔍 Collecting parameters from '{SHEET_NAME}' sheet...")

for row in range(2, param_ws.max_row + 1):
    param_value = param_ws.cell(row=row, column=param_col_idx).value
    readonly_value = param_ws.cell(row=row, column=readonly_col_idx).value

    if param_value is not None and str(param_value).strip() != "":
        param_clean = str(param_value).strip()

        # Skip administrativeState parameter
        if "administrativestate" in param_clean.lower():
            print(f"⚠️  Skipping administrativeState parameter: {param_clean}")
            continue

        if "nrtac" in param_clean.lower():
            print(f"⚠️  Skipping nRTAC parameter: {param_clean}")
            continue

        # Skip excluded parameters
        if param_clean in EXCLUDED_PARAMETERS:
            print(f"⚠️  Skipping excluded parameter: {param_clean}")
            continue

        # Check if parameter is NOT read-only
        if readonly_value is None or str(readonly_value).strip().lower() != "read-only":
            parameters_to_include.append(param_clean)

            # Store parameter data for later use
            parameter_data[param_clean] = {
                "Valeur par défaut RBS": param_ws.cell(row=row,
                                                       column=param_col_idx_map["Valeur par défaut RBS"]).value,
                "Valeur Bytel TDD MidBand": param_ws.cell(row=row,
                                                          column=param_col_idx_map["Valeur Bytel TDD MidBand"]).value,
                "Valeur Bytel FDD ESS 15MHz": param_ws.cell(row=row, column=param_col_idx_map[
                    "Valeur Bytel FDD ESS 15MHz"]).value,
                "Valeur Bytel TDD HigBand": param_ws.cell(row=row,
                                                          column=param_col_idx_map["Valeur Bytel TDD HigBand"]).value
            }
            
            # Add TDD+FDD co-node value for NRCellCU
            if SHEET_NAME == "NRCellCU":
                parameter_data[param_clean]["Valeur Bytel TDD+FDD co-node"] = param_ws.cell(row=row,
                    column=param_col_idx_map["Valeur Bytel TDD+FDD co-node\nAppliquer la valeur commune si valeur TDD et FDD sont même, sinon appliquer la valeur spécifiée dans cette colonne."]).value


print(f"✅ Collected {len(parameters_to_include)} parameters (excluding read-only and administrativeState and nRTAC)")

# Load data workbook
try:
    print(f"\n📖 Loading data workbook...")
    data_df = pd.read_excel(data_path, engine="openpyxl")
                
except Exception as e:
    print(f"❌ Error loading data file: {e}")
    raise SystemExit


# Check if data file has required columns
if "CellName" not in data_df.columns:
    print("❌ 'CellName' column not found in data file")
    raise SystemExit

# Check if NeName column exists for node type categorization
if "NeName" not in data_df.columns:
    print("❌ 'NeName' column not found in data file - needed for node type categorization")
    raise SystemExit

# Find which parameters from our list exist in the data file
available_parameters_in_data = []
for param in parameters_to_include:
    if param in data_df.columns:
        available_parameters_in_data.append(param)
    else:
        missing_parameters_in_data.append(param)

print(f"🔍 Found {len(available_parameters_in_data)} parameters in data file")
print(f"❌ {len(missing_parameters_in_data)} parameters not found in data file")

if not available_parameters_in_data:
    print("❌ No matching parameters found between parameter file and data file")
    raise SystemExit

# Create DataFrames for output
main_output_data = []
wrong_parameters_data = []  # For incorrect values
missing_data_records = []  # For records with no data

print(f"\n📝 Creating structured output table with validation...")

# Define color fills
GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light Green
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
VIOLET_FILL = PatternFill(start_color="CBC3E3", end_color="CBC3E3", fill_type="solid")  # Light Purple/Violet

validation_stats = {
    "correct": 0,
    "correct_fuzzy": 0,
    "correct_special": 0,
    "correct_extracted": 0,
    "correct_numeric": 0,
    "incorrect": 0,
    "no_data": 0,
    "no_expected_value": 0,
    "skipped": 0
}

# Process EACH parameter and EACH row individually
for param in available_parameters_in_data:
    # Get parameter info from parameter file
    param_info = parameter_data[param]

    # Process EACH row individually
    for idx, row in data_df.iterrows():
        value = row[param]
        cellname = row['CellName']
        nename = row['NeName']
        
        # Skip if all key fields are empty
        if pd.isna(value) and pd.isna(cellname) and pd.isna(nename):
            continue
            
        cell_type = get_cell_type(cellname)
        node_type = get_node_type(nename, nename_categories)
        
        # Get category information for this NeName
        category_info = nename_categories.get(str(nename).strip(), {})
        
        # Use original value without conversion
        display_value = convert_for_display(value)
        
        # Get co-node value for NRCellCU
        expected_co_node_value = None
        if SHEET_NAME == "NRCellCU":
            expected_co_node_value = param_info.get("Valeur Bytel TDD+FDD co-node")
        
        # Get Gen value for Gen2 handling
        gen_value = category_info.get('Gen', '')
        
        # Get operator from cell name (last letter)
        operateur = get_operator_from_cellname(cellname)
        
        validation = validate_parameter_value(
            value,
            param_info["Valeur Bytel TDD MidBand"] or param_info["Valeur Bytel TDD HigBand"],
            param_info["Valeur Bytel FDD ESS 15MHz"],
            param_info["Valeur par défaut RBS"],
            expected_co_node_value,
            cell_type,
            param,
            node_type,
            operateur,  # Now using operator from cell name
            category_info.get('Remarque', ''),
            cellname,
            gen_value
        )

        validation_stats[validation] += 1

        # Track missing data ONLY if it's truly missing (not empty by design)
        if validation == "no_data":
            missing_data_records.append({
                "Parameter": param,
                "CellName": cellname,
                "NeName": nename,
                "CellType": cell_type,
                "Type": category_info.get('Type', ''),
                "Operateur": operateur,  # Using operator from cell name
                "Cell": category_info.get('Cell', ''),
                "Gen": category_info.get('Gen', ''),
                "Remarque": category_info.get('Remarque', ''),
                "Valeur par défaut RBS": param_info["Valeur par défaut RBS"],
                "Valeur Bytel TDD MidBand": param_info["Valeur Bytel TDD MidBand"],
                "Valeur Bytel FDD ESS 15MHz": param_info["Valeur Bytel FDD ESS 15MHz"],
                "Valeur Bytel TDD HigBand": param_info["Valeur Bytel TDD HigBand"]
            })

        # Add to main output
        output_row = {
            "Parameter": param,
            "Valeur par défaut RBS": param_info["Valeur par défaut RBS"],
            "Valeur Bytel TDD MidBand": param_info["Valeur Bytel TDD MidBand"],
            "Valeur Bytel FDD ESS 15MHz": param_info["Valeur Bytel FDD ESS 15MHz"],
            "Valeur Bytel TDD HigBand": param_info["Valeur Bytel TDD HigBand"],
            "Value": display_value,
            "CellName": cellname,
            "NeName": nename,
            "CellType": cell_type,
            "Type": category_info.get('Type', ''),
            "Operateur": operateur,  # Using operator from cell name
            "Cell": category_info.get('Cell', ''),
            "Gen": category_info.get('Gen', ''),
            "Remarque": category_info.get('Remarque', ''),
            "Validation": validation
        }
        
        # Add co-node value for NRCellCU
        if SHEET_NAME == "NRCellCU":
            output_row["Valeur Bytel TDD+FDD co-node"] = expected_co_node_value
            
        main_output_data.append(output_row)

        # Add to wrong parameters if incorrect
        if validation == "incorrect":
            wrong_row = {
                "Parameter": param,
                "Valeur par défaut RBS": param_info["Valeur par défaut RBS"],
                "Valeur Bytel TDD MidBand": param_info["Valeur Bytel TDD MidBand"],
                "Valeur Bytel FDD ESS 15MHz": param_info["Valeur Bytel FDD ESS 15MHz"],
                "Valeur Bytel TDD HigBand": param_info["Valeur Bytel TDD HigBand"],
                "Actual_Value": convert_for_display(value),
                "Expected_Value": get_display_expected_value(param_info, cell_type, node_type, expected_co_node_value, param, cellname, operateur),
                "CellName": cellname,
                "NeName": nename,
                "CellType": cell_type,
                "Type": category_info.get('Type', ''),
                "Operateur": operateur,  # Using operator from cell name
                "Cell": category_info.get('Cell', ''),
                "Gen": category_info.get('Gen', ''),
                "Remarque": category_info.get('Remarque', '')
            }
            
            # Add co-node value for NRCellCU
            if SHEET_NAME == "NRCellCU":
                wrong_row["Valeur Bytel TDD+FDD co-node"] = expected_co_node_value
                
            wrong_parameters_data.append(wrong_row)

# Create output DataFrames
main_output_df = pd.DataFrame(main_output_data)
wrong_parameters_df = pd.DataFrame(wrong_parameters_data)
missing_data_df = pd.DataFrame(missing_data_records)

# Create summary data for the Summary sheet
summary_data = {
    "Category": [
        "Total Parameters in Parameter File",
        "Parameters Found in Data File",
        "Parameters Not Found in Data File",
        "Total Validations Performed",
        "Correct Values (Exact Match)",
        "Correct Values (Fuzzy Match)",
        "Correct Values (Special Rules)",
        "Correct Values (Extracted)",
        "Correct Values (Numeric)",
        "Incorrect Values",
        "Missing Data",
        "No Expected Value",
        "Skipped (administrativeState and nRTAC)"
    ],
    "Count": [
        len(parameters_to_include),
        len(available_parameters_in_data),
        len(missing_parameters_in_data),
        sum(validation_stats.values()),
        validation_stats["correct"],
        validation_stats["correct_fuzzy"],
        validation_stats["correct_special"],
        validation_stats["correct_extracted"],
        validation_stats["correct_numeric"],
        validation_stats["incorrect"],
        validation_stats["no_data"],
        validation_stats["no_expected_value"],
        validation_stats["skipped"]
    ]
}

summary_df = pd.DataFrame(summary_data)

print(f"✅ Created main output table with {len(main_output_df)} rows")
print(f"✅ Created wrong parameters table with {len(wrong_parameters_df)} rows")
print(f"✅ Created missing data table with {len(missing_data_df)} rows")
print(f"📊 Validation Statistics:")
print(f"   ⚠️  Incorrect values: {validation_stats['incorrect']}")
print(f"   💜 No data: {validation_stats['no_data']}")
print(f"   ❓ No expected value: {validation_stats['no_expected_value']}")
print(f"   ⏭️  Skipped (administrativeState and nRTAC): {validation_stats['skipped']}")

# Show parameters not found in data file
if missing_parameters_in_data:
    print(f"\n❌ Parameters not found in data file ({len(missing_parameters_in_data)}):")
    for param in missing_parameters_in_data:
        print(f"   - {param}")

# Create output file name
output_filename = f"{SHEET_NAME}_Parameter_Validation.xlsx"

# Save to new Excel file with multiple sheets
try:
    print(f"\n💾 Saving output to: {output_filename}")

    # Create a new workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create main validation sheet
    ws_main = wb.create_sheet("Parameter_Validation")

    # Write headers for main sheet
    main_headers = ["Parameter", "Valeur par défaut RBS", "Valeur Bytel TDD MidBand",
                    "Valeur Bytel FDD ESS 15MHz", "Valeur Bytel TDD HigBand", "Value",
                    "CellName", "NeName", "CellType", "Type", "Operateur", 
                    "Cell", "Gen", "Remarque", "Validation"]
    
    # Add TDD+FDD co-node column for NRCellCU
    if SHEET_NAME == "NRCellCU":
        main_headers.insert(5, "Valeur Bytel TDD+FDD co-node")

    for col_idx, header in enumerate(main_headers, 1):
        ws_main.cell(row=1, column=col_idx, value=header)

    # Write main data and apply formatting
    for row_idx, (_, row_data) in enumerate(main_output_df.iterrows(), 2):
        # Write row data
        for col_idx, header in enumerate(main_headers, 1):
            ws_main.cell(row=row_idx, column=col_idx, value=row_data[header])

        # Apply color based on validation
        validation = row_data["Validation"]
        if validation in ["correct", "correct_fuzzy", "correct_special", "correct_extracted", "correct_numeric"]:
            fill_color = GREEN_FILL
        elif validation == "incorrect":
            fill_color = YELLOW_FILL
        elif validation == "no_data":
            fill_color = VIOLET_FILL
        elif validation == "skipped":
            fill_color = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light Pink
        else:
            fill_color = None

        if fill_color:
            # Apply color to Value and CellName columns
            value_col = 6 if SHEET_NAME == "NRCellDU" else 7  # Adjust column index for co-node
            cellname_col = value_col + 1
            ws_main.cell(row=row_idx, column=value_col).fill = fill_color
            ws_main.cell(row=row_idx, column=cellname_col).fill = fill_color

    # Create wrong parameters sheet
    if len(wrong_parameters_df) > 0:
        ws_wrong = wb.create_sheet("Wrong_Parameters")

        # Write headers for wrong parameters sheet
        wrong_headers = ["Parameter", "Valeur par défaut RBS", "Valeur Bytel TDD MidBand",
                         "Valeur Bytel FDD ESS 15MHz", "Valeur Bytel TDD HigBand",
                         "Actual_Value", "Expected_Value", "CellName", "NeName", "CellType", 
                         "Type", "Operateur", "Cell", "Gen", "Remarque"]
        
        # Add TDD+FDD co-node column for NRCellCU
        if SHEET_NAME == "NRCellCU":
            wrong_headers.insert(5, "Valeur Bytel TDD+FDD co-node")

        for col_idx, header in enumerate(wrong_headers, 1):
            ws_wrong.cell(row=1, column=col_idx, value=header)

        # Write wrong parameters data
        for row_idx, (_, row_data) in enumerate(wrong_parameters_df.iterrows(), 2):
            for col_idx, header in enumerate(wrong_headers, 1):
                ws_wrong.cell(row=row_idx, column=col_idx, value=row_data[header])

            # Highlight the incorrect values in yellow
            actual_value_col = 6 if SHEET_NAME == "NRCellDU" else 7
            expected_value_col = actual_value_col + 1
            ws_wrong.cell(row=row_idx, column=actual_value_col).fill = YELLOW_FILL
            ws_wrong.cell(row=row_idx, column=expected_value_col).fill = YELLOW_FILL

    # Create missing data sheet only if there are missing data records
    if len(missing_data_df) > 0:
        ws_missing = wb.create_sheet("Missing_Data")

        # Write headers for missing data sheet
        missing_headers = ["Parameter", "CellName", "NeName", "CellType", "Type", 
                           "Operateur", "Cell", "Gen", "Remarque", "Valeur par défaut RBS",
                           "Valeur Bytel TDD MidBand", "Valeur Bytel FDD ESS 15MHz", "Valeur Bytel TDD HigBand"]

        for col_idx, header in enumerate(missing_headers, 1):
            ws_missing.cell(row=1, column=col_idx, value=header)

        # Write missing data
        for row_idx, (_, row_data) in enumerate(missing_data_df.iterrows(), 2):
            for col_idx, header in enumerate(missing_headers, 1):
                ws_missing.cell(row=row_idx, column=col_idx, value=row_data[header])

            # Highlight in violet to indicate missing data
            for col in range(1, len(missing_headers) + 1):
                ws_missing.cell(row=row_idx, column=col).fill = VIOLET_FILL

    # Create summary sheet
    ws_summary = wb.create_sheet("Summary")

    # Write summary headers
    ws_summary.cell(row=1, column=1, value="Validation Summary").font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:B1')

    # Write summary data
    for row_idx, (_, row_data) in enumerate(summary_df.iterrows(), 3):
        ws_summary.cell(row=row_idx, column=1, value=row_data["Category"])
        ws_summary.cell(row=row_idx, column=2, value=row_data["Count"])

    # Add parameters not found section
    start_row = len(summary_df) + 5
    ws_summary.cell(row=start_row, column=1, value="Parameters Not Found in Data File").font = Font(bold=True)
    ws_summary.merge_cells(f'A{start_row}:B{start_row}')

    if missing_parameters_in_data:
        for i, param in enumerate(missing_parameters_in_data, start_row + 1):
            ws_summary.cell(row=i, column=1, value=param)
    else:
        ws_summary.cell(row=start_row + 1, column=1, value="All parameters were found in data file")

    # Auto-adjust column widths for all sheets
    sheets_to_adjust = [ws_main, ws_summary]
    if len(wrong_parameters_df) > 0:
        sheets_to_adjust.append(ws_wrong)
    if len(missing_data_df) > 0:
        sheets_to_adjust.append(ws_missing)

    for ws in sheets_to_adjust:
        for column_cells in ws.columns:
            # Skip if it's a MergedCell (which doesn't have column_letter attribute)
            if isinstance(column_cells[0], MergedCell):
                continue

            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    # Skip MergedCell objects
                    if isinstance(cell, MergedCell):
                        continue
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Add borders and formatting to all sheets
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for ws in sheets_to_adjust:
        if ws == ws_main:
            max_row = len(main_output_df) + 1
            max_col = len(main_headers)
        elif ws == ws_wrong and len(wrong_parameters_df) > 0:
            max_row = len(wrong_parameters_df) + 1
            max_col = len(wrong_headers)
        elif len(missing_data_df) > 0:
            max_row = len(missing_data_df) + 1
            max_col = len(missing_headers)
        elif ws == ws_summary:
            max_row = start_row + len(missing_parameters_in_data) if missing_parameters_in_data else start_row + 1
            max_col = 2
        else:
            continue

        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                # Only apply border to non-merged cells
                if not isinstance(cell, MergedCell):
                    cell.border = thin_border

        # Style header row (skip merged cells)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            if not isinstance(cell, MergedCell):
                cell.fill = header_fill
                cell.font = Font(bold=True)

    # Add legend to main sheet
    legend_row = len(main_output_df) + 3
    ws_main.cell(row=legend_row, column=1, value="LEGEND:").font = Font(bold=True)
    ws_main.cell(row=legend_row + 1, column=1, value="Green").fill = GREEN_FILL
    ws_main.cell(row=legend_row + 1, column=2, value="= Correct value")
    ws_main.cell(row=legend_row + 2, column=1, value="Yellow").fill = YELLOW_FILL
    ws_main.cell(row=legend_row + 2, column=2, value="= Incorrect value")
    ws_main.cell(row=legend_row + 3, column=1, value="Violet").fill = VIOLET_FILL
    ws_main.cell(row=legend_row + 3, column=2, value="= No data")
    ws_main.cell(row=legend_row + 4, column=1, value="Pink").fill = PatternFill(start_color="FFB6C1",
                                                                                end_color="FFB6C1", fill_type="solid")
    ws_main.cell(row=legend_row + 4, column=2, value="= Skipped (administrativeState and nRTAC)")

    wb.save(output_filename)

    print(f"🎉 SUCCESS!")
    print(f"📁 Output file created: {output_filename}")
    print(f"📊 Sheets created:")
    print(f"   - Parameter_Validation: Main validation results")
    if len(wrong_parameters_df) > 0:
        print(f"   - Wrong_Parameters: {len(wrong_parameters_df)} incorrect values")
    if len(missing_data_df) > 0:
        print(f"   - Missing_Data: {len(missing_data_df)} missing data records")
    print(f"   - Summary: Overall validation summary")

except Exception as e:
    print(f"❌ Error saving output file: {e}")
    import traceback
    traceback.print_exc()
