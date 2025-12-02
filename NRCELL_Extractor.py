import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def check_cell_types_in_file(input_file_path):
    """Check if both NRCellDU and NRCellCU sections exist in the file"""
    with open(input_file_path, 'r') as file:
        content = file.read()
    
    has_nrcelldu = "NRCellDU=" in content
    has_nrcellcu = "NRCellCU=" in content
    
    return has_nrcelldu, has_nrcellcu

def extract_parameters(input_file_path, cell_type):
    with open(input_file_path, 'r') as file:
        lines = file.readlines()

    found_section = False
    parameter_names = []

    current_parent_parameter = ""

    for line in lines:
        section_identifier = f"{cell_type}="

        if section_identifier in line:
            found_section = True
            continue

        if found_section:
            stripped_line = line.strip()

            # Check if we've reached next cell type section
            if stripped_line.startswith("NRCellDU=") and cell_type != "NRCellDU":
                break
            if stripped_line.startswith("NRCellCU=") and cell_type != "NRCellCU":
                break

            if (
                stripped_line == "" or
                stripped_line.startswith(("===", "Total:", "INFO:")) or
                stripped_line[0].isdigit() or
                stripped_line[0] in {'X', 'G', 'E'}
            ):
                continue

            if stripped_line.startswith(">>>"):
                parts = stripped_line.split('.')
                if len(parts) > 1:
                    sub_parameter = parts[1].split('=')[0].strip()
                    parameter_names.append((current_parent_parameter,
sub_parameter))
            else:
                current_parent_parameter = stripped_line.split()[0].strip()
                parameter_names.append((None, current_parent_parameter))

    # Create DataFrame for the extracted parameters
    extracted_data = []
    for parent, sub in parameter_names:
        if parent:
            extracted_data.append({'Struct': parent, 'Parameter': sub})
        else:
            extracted_data.append({'Struct': '', 'Parameter': sub})

    return pd.DataFrame(extracted_data)

def normalize_parameter_name(param):
    """Normalize parameter name by removing extra spaces and
standardizing format"""
    if pd.isna(param):
        return ""
    # Remove extra spaces and normalize
    param = str(param).strip()
    # Replace multiple spaces with single space
    param = ' '.join(param.split())
    return param

def detect_barred_parameters_from_excel(excel_file_path, filtered_df, param_col_idx, sheet_name="LTE - NR parameters"):
    """Detect parameters with strikethrough formatting in Excel"""
    try:
        # Load workbook with openpyxl to check formatting
        workbook = load_workbook(excel_file_path, data_only=True)
        
        # Try to find the correct sheet
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            # Try to find sheet with similar name
            for ws_name in workbook.sheetnames:
                if "LTE" in ws_name or "NR" in ws_name or "param" in ws_name.lower():
                    sheet = workbook[ws_name]
                    break
            else:
                sheet = workbook[workbook.sheetnames[0]]
        
        # Find header row to get column indices
        header_row = None
        for row in range(1, 20):  # Check first 20 rows
            cell_value = sheet.cell(row=row, column=param_col_idx).value
            if cell_value and "parameter" in str(cell_value).lower():
                header_row = row
                break
        
        if not header_row:
            # Default to row 1
            header_row = 1
        
        barred_parameters = []
        barred_indices = []
        
        # Check each row in filtered_df for strikethrough
        for idx, row in filtered_df.iterrows():
            excel_row = idx + header_row + 1  # Convert DataFrame index to Excel row
            
            # Get the cell in the parameter column
            param_cell = sheet.cell(row=excel_row, column=param_col_idx)
            
            # Check if cell has strikethrough formatting
            if param_cell.font and param_cell.font.strike:
                param_name = str(param_cell.value).strip() if param_cell.value else ""
                if param_name:
                    barred_parameters.append(param_name)
                    barred_indices.append(idx)
        
        return barred_parameters, barred_indices
        
    except Exception as e:
        print(f"Warning: Could not detect strikethrough formatting: {e}")
        return [], []

def compare_with_excel_sheet(lab_params_df, excel_file_path,
cell_type, output_file_path):
    # Read parameters from LAB DataFrame
    lab_params = set()
    lab_params_dict = {}  # Store original parameter with struct info

    for _, row in lab_params_df.iterrows():
        struct = row['Struct'].strip() if pd.notna(row['Struct']) else ""
        param = row['Parameter'].strip() if pd.notna(row['Parameter']) else ""
        normalized_param = normalize_parameter_name(param)

        if struct:
            full_param = f"{struct}.{normalized_param}"
            lab_params.add(full_param)
            lab_params_dict[full_param] = {'struct': struct,
'subparam': normalized_param}
        else:
            lab_params.add(normalized_param)
            lab_params_dict[normalized_param] = {'struct': '',
'subparam': normalized_param}

    # Read Excel file
    SHEET_NAME = "LTE - NR parameters"
    PARAM_COL_NAME = "Parameter"
    MO_COL_NAME = "MO"

    try:
        # Read the Excel file
        df = pd.read_excel(excel_file_path, sheet_name=SHEET_NAME)

        # More precise filtering based on cell type
        if cell_type == "NRCellDU":
            # For NRCellDU, include all variations
            filtered_df = df[df[MO_COL_NAME].str.contains(r'^NRCellDU', na=False)]
        elif cell_type == "NRCellCU":
            # For NRCellCU, only exact NRCellCU (not containing commas for other MOs)
            filtered_df = df[df[MO_COL_NAME] == "NRCellCU"]
        else:
            # For other cell types, use exact match
            filtered_df = df[df[MO_COL_NAME] == cell_type]

        print(f"Found {len(filtered_df)} parameters for {cell_type} in Excel sheet")

        # STEP 1: Track all parameters before any filtering
        all_parameters_step1 = []
        for idx, row in filtered_df.iterrows():
            original_param = str(row[PARAM_COL_NAME]) if pd.notna(row[PARAM_COL_NAME]) else ""
            if original_param and original_param.strip():
                all_parameters_step1.append({
                    'original': original_param.strip(),
                    'row_index': idx
                })

        print(f"Step 1 - Initial parameters: {len(all_parameters_step1)}")

        # STEP 2: Detect barred parameters with strikethrough formatting
        # Find parameter column index
        param_col_idx = df.columns.get_loc(PARAM_COL_NAME) + 1  # Convert to 1-based index
        
        barred_parameters, barred_indices = detect_barred_parameters_from_excel(
            excel_file_path, filtered_df, param_col_idx, SHEET_NAME
        )
        
        # Remove barred parameters from filtered_df
        if barred_parameters:
            print(f"Found {len(barred_parameters)} parameters with strikethrough formatting (barred/obsolete):")
            for i, param in enumerate(barred_parameters, 1):
                print(f"  {i}. {param}")
            
            # Remove barred rows
            filtered_df = filtered_df.drop(barred_indices)

        # STEP 3: Track parameters after removing barred
        all_parameters_step2 = []
        for idx, row in filtered_df.iterrows():
            original_param = str(row[PARAM_COL_NAME]) if pd.notna(row[PARAM_COL_NAME]) else ""
            if original_param and original_param.strip():
                all_parameters_step2.append({
                    'original': original_param.strip(),
                    'row_index': idx
                })

        print(f"Step 2 - After removing barred: {len(all_parameters_step2)}")

        # STEP 4: Normalize and check for duplicates
        normalized_params_dict = {}
        duplicate_normalized = []
        empty_params = []
        
        for param_info in all_parameters_step2:
            original = param_info['original']
            normalized = normalize_parameter_name(original)
            
            if not normalized or normalized == "" or normalized.lower() == "nan":
                empty_params.append(original)
                continue
            
            if normalized in normalized_params_dict:
                duplicate_normalized.append({
                    'original': original,
                    'normalized': normalized,
                    'first_occurrence': normalized_params_dict[normalized]
                })
            else:
                normalized_params_dict[normalized] = original

        # Remove empty parameters
        if empty_params:
            print(f"Removed {len(empty_params)} empty parameters")
        
        # Remove duplicate normalized parameters (keep first occurrence)
        if duplicate_normalized:
            print(f"Removed {len(duplicate_normalized)} duplicate parameters after normalization")

        # Final expected parameters
        expected_params = set(normalized_params_dict.keys())
        
        # STEP 5: Find ALL parameters that were filtered out
        all_filtered_out = []
        
        # Add barred parameters
        for barred in barred_parameters:
            all_filtered_out.append(f"{barred} (barred - strikethrough formatting)")
        
        # Add empty parameters
        for empty in empty_params:
            all_filtered_out.append(f"{empty} (empty)")
        
        # Add duplicate normalized parameters (except the first occurrence)
        for dup in duplicate_normalized:
            all_filtered_out.append(f"{dup['original']} (duplicate of: {dup['first_occurrence']})")
        
        # Show filtered out parameters
        if all_filtered_out:
            print(f"\nPARAMETERS FILTERED OUT (not included in comparison):")
            print(f"Total filtered out: {len(all_filtered_out)}")
            for i, param in enumerate(sorted(set(all_filtered_out)), 1):
                print(f"  {i}. {param}")
        else:
            print(f"\nNo parameters were filtered out")

        # Find matching and missing parameters
        matching_params = lab_params.intersection(expected_params)
        missing_params = expected_params - lab_params

        # Create comparison DataFrame with clear column names
        comparison_data = []

        for param in sorted(matching_params):
            lab_info = lab_params_dict.get(param, {'struct': '', 'subparam': param})
            comparison_data.append({
                'Parameter_in_Excel': normalized_params_dict.get(param, param),
                'Parameter_in_LAB': param,
                'Struct_in_LAB': lab_info['struct'],
                'Status': 'PRESENT'
            })

        for param in sorted(missing_params):
            comparison_data.append({
                'Parameter_in_Excel': normalized_params_dict.get(param, param),
                'Parameter_in_LAB': 'NOT FOUND',
                'Struct_in_LAB': '',
                'Status': 'MISSING'
            })

        # Create DataFrame for comparison
        comparison_df = pd.DataFrame(comparison_data)

        # Reorder columns to have Status as the last column
        column_order = ['Parameter_in_Excel', 'Parameter_in_LAB', 'Struct_in_LAB', 'Status']
        comparison_df = comparison_df[column_order]

        # Create summary DataFrame
        summary_data = {
            'Category': ['Total Parameters in Excel Sheet', 
                        'Parameters with Strikethrough (Barred)', 
                        'Empty/Duplicate Parameters Removed', 
                        'Total Expected Parameters', 
                        'Parameters Found in LAB', 
                        'Missing Parameters'],
            'Count': [len(all_parameters_step1), 
                     len(barred_parameters), 
                     len(empty_params) + len(duplicate_normalized),
                     len(expected_params), 
                     len(matching_params), 
                     len(missing_params)]
        }
        summary_df = pd.DataFrame(summary_data)

        # Save all sheets to the same Excel file
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            # Sheet 1: Extracted parameters from LAB
            lab_params_df.to_excel(writer, sheet_name='Extracted Parameters', index=False)

            # Sheet 2: Summary
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

            # Sheet 3: Detailed comparison
            comparison_df.to_excel(writer, sheet_name='Detailed Comparison', index=False)

        # Apply formatting to highlight missing parameters in yellow
        apply_formatting_to_excel(output_file_path)

        print(f"\nComparison completed successfully!")
        print(f"Total parameters in Excel sheet: {len(all_parameters_step1)}")
        print(f"Parameters with strikethrough (barred): {len(barred_parameters)}")
        print(f"Empty/duplicate parameters removed: {len(empty_params) + len(duplicate_normalized)}")
        print(f"Total expected parameters after filtering: {len(expected_params)}")
        print(f"Parameters Excel found in LAB: {len(matching_params)}")
        print(f"Missing parameters: {len(missing_params)}")
        print(f"All results saved to: {output_file_path}")
        
        # Return both success status and missing parameters list
        missing_params_list = []
        for param in sorted(missing_params):
            missing_params_list.append(normalized_params_dict.get(param, param))
        
        return len(missing_params) == 0, missing_params_list  # Return True if no missing parameters

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False, []

def apply_formatting_to_excel(file_path):
    """Apply yellow highlighting to missing parameters in the Detailed
Comparison sheet"""
    try:
        workbook = load_workbook(file_path)

        if 'Detailed Comparison' in workbook.sheetnames:
            sheet = workbook['Detailed Comparison']

            # Yellow fill for missing parameters
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            # Find the Status column
            header_row = 1
            status_col_idx = None

            for col_idx, cell in enumerate(sheet[header_row], 1):
                if cell.value == 'Status':
                    status_col_idx = col_idx
                    break

            if status_col_idx:
                # Apply formatting to all rows
                for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 (after header)
                    status_cell = sheet.cell(row=row_idx, column=status_col_idx)
                    if status_cell.value == 'MISSING':
                        # Highlight the entire row for missing parameters
                        for col_idx in range(1, sheet.max_column + 1):
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            cell.fill = yellow_fill

        workbook.save(file_path)
        print("Formatting applied successfully - missing parameters highlighted in yellow")

    except Exception as e:
        print(f"Error applying formatting: {e}")

def main():
    input_file_path = input("Please enter the path to the input LAB file: ").strip()
    
    # Check if both cell types exist in the file
    print("\n" + "="*60)
    print("Checking input file for NRCellDU and NRCellCU sections...")
    has_nrcelldu, has_nrcellcu = check_cell_types_in_file(input_file_path)
    
    if not has_nrcelldu and not has_nrcellcu:
        print("ERROR: The input file does not contain either NRCellDU= or NRCellCU= sections!")
        print("Please provide a valid LAB file with at least one of these sections.")
        return
    elif not has_nrcelldu:
        print("ERROR: The input file does not contain NRCellDU= section!")
        print("Please provide a LAB file with NRCellDU= section.")
        return
    elif not has_nrcellcu:
        print("ERROR: The input file does not contain NRCellCU= section!")
        print("Please provide a LAB file with NRCellCU= section.")
        return
    
    print("✓ File contains both NRCellDU and NRCellCU sections")
    print("="*60)
    
    # Extract parameters for both cell types
    print("\nExtracting parameters for both NRCellDU and NRCellCU...")
    
    nrcelldu_params_df = extract_parameters(input_file_path, "NRCellDU")
    nrcellcu_params_df = extract_parameters(input_file_path, "NRCellCU")
    
    print(f"\n✓ Extracted {len(nrcelldu_params_df)} parameters for NRCellDU")
    print(f"✓ Extracted {len(nrcellcu_params_df)} parameters for NRCellCU")
    
    # Create output files for each cell type
    script_directory = os.path.dirname(__file__)
    
    # Process NRCellDU
    print("\n" + "="*60)
    print("Processing NRCellDU parameters...")
    print("="*60)
    
    nrcelldu_output_file = f"Parameter_Analysis_NRCellDU.xlsx"
    nrcelldu_output_path = os.path.join(script_directory, nrcelldu_output_file)
    
    # Save extracted parameters to the Excel file first
    with pd.ExcelWriter(nrcelldu_output_path, engine='openpyxl') as writer:
        nrcelldu_params_df.to_excel(writer, sheet_name='Extracted Parameters in LAB', index=False)

    print(f"NRCellDU parameters extracted and saved to: {nrcelldu_output_path}")
    
    # Process NRCellCU
    print("\n" + "="*60)
    print("Processing NRCellCU parameters...")
    print("="*60)
    
    nrcellcu_output_file = f"Parameter_Analysis_NRCellCU.xlsx"
    nrcellcu_output_path = os.path.join(script_directory, nrcellcu_output_file)
    
    # Save extracted parameters to the Excel file first
    with pd.ExcelWriter(nrcellcu_output_path, engine='openpyxl') as writer:
        nrcellcu_params_df.to_excel(writer, sheet_name='Extracted Parameters in LAB', index=False)

    print(f"NRCellCU parameters extracted and saved to: {nrcellcu_output_path}")
    
    # Ask if user wants to compare with Excel sheet
    compare_choice = input("\nDo you want to compare with an Excel parameter sheet? (y/n): ").strip().lower()

    all_missing_params = []
    
    if compare_choice in ['y', 'yes']:
        excel_file_path = input("Please enter the path to the Excel parameter sheet: ").strip()

        # Compare NRCellDU
        print("\n" + "="*60)
        print("Comparing NRCellDU parameters with Excel sheet...")
        print("="*60)
        success_du, missing_du = compare_with_excel_sheet(nrcelldu_params_df, excel_file_path, "NRCellDU", nrcelldu_output_path)
        
        # Compare NRCellCU
        print("\n" + "="*60)
        print("Comparing NRCellCU parameters with Excel sheet...")
        print("="*60)
        success_cu, missing_cu = compare_with_excel_sheet(nrcellcu_params_df, excel_file_path, "NRCellCU", nrcellcu_output_path)
        
        # Collect all missing parameters
        all_missing_params = missing_du + missing_cu
        
        print("\n" + "="*60)
        print("SUMMARY OF MISSING PARAMETERS:")
        print("="*60)
        
        if missing_du:
            print(f"\nMISSING NRCellDU PARAMETERS ({len(missing_du)}):")
            for i, param in enumerate(missing_du, 1):
                print(f"{i:3}. {param}")
        
        if missing_cu:
            print(f"\nMISSING NRCellCU PARAMETERS ({len(missing_cu)}):")
            for i, param in enumerate(missing_cu, 1):
                print(f"{i:3}. {param}")
        
        if not missing_du and not missing_cu:
            print("\n✓ Excellent! No missing parameters found for either cell type!")
        
        print("="*60)
        
        if success_du and success_cu:
            print("\n✓ Both comparisons completed successfully!")
        else:
            print("\n⚠ Comparisons completed with some missing parameters")
            
    else:
        print(f"\nOnly extracted parameters saved to separate files.")
        print(f"NRCellDU: {nrcelldu_output_path}")
        print(f"NRCellCU: {nrcellcu_output_path}")

if __name__ == "__main__":
    main()
