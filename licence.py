import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.cell.cell import MergedCell
import zipfile
import re


def try_open_excel(path):
    candidates = [path]
    if not os.path.splitext(path)[1]:
        candidates += [path + ext for ext in (".xlsx", ".xls", ".xlsm")]
    for p in candidates:
        if os.path.exists(p):
            try:
                if p.endswith(('.xlsx', '.xlsm')):
                    with zipfile.ZipFile(p, 'r') as zf:
                        pass
                return p
            except (zipfile.BadZipFile, Exception):
                continue
    return None


def is_valid_excel_file(file_path):
    """Check if the file is a valid Excel file"""
    try:
        if file_path.endswith(('.xlsx', '.xlsm')):
            with zipfile.ZipFile(file_path, 'r') as zf:
                return True
        elif file_path.endswith('.xls'):
            pd.read_excel(file_path, nrows=1, engine='xlrd')
            return True
        else:
            pd.read_excel(file_path, nrows=1)
            return True
    except Exception as e:
        print(f"❌ File validation failed for {file_path}: {e}")
        return False


def get_node_type(node_name):
    """Extract node type from NeName (E, X, G)"""
    if node_name == "NOT FOUND":
        return None
    
    if isinstance(node_name, str):
        if node_name.startswith('E'):
            return 'E'
        elif node_name.startswith('X'):
            return 'X'
        elif node_name.startswith('G'):
            return 'G'
    return None

def load_nename_categories(category_file_path):
    """Load NeName categories from Excel file"""
    try:
        print(f"📖 Loading NeName categories from: {category_file_path}")
        category_df = pd.read_excel(category_file_path, engine='openpyxl')
        
        # Debug: Print available columns
        print(f"📋 Available columns in category file: {list(category_df.columns)}")
        
        # Create a mapping dictionary for NeName to categories
        nename_categories = {}
        
        # Check if required columns exist
        required_columns = ['NeName', 'Type', 'Operateur', 'Cell', 'Gen', 'Remarque']
        missing_columns = [col for col in required_columns if col not in category_df.columns]
        
        if missing_columns:
            print(f"❌ Missing columns in category file: {missing_columns}")
            print(f"✅ Available columns: {list(category_df.columns)}")
            return {}
        
        # Check if dual-co column exists
        dual_co_column = None
        for col in category_df.columns:
            if 'dual' in str(col).lower() and 'co' in str(col).lower():
                dual_co_column = col
                break
        
        for _, row in category_df.iterrows():
            if pd.isna(row['NeName']):
                continue
                
            nename = str(row['NeName']).strip()
            category_info = {
                'Type': row['Type'] if pd.notna(row['Type']) else '',
                'Operateur': row['Operateur'] if pd.notna(row['Operateur']) else '',
                'Cell': row['Cell'] if pd.notna(row['Cell']) else '',
                'Gen': row['Gen'] if pd.notna(row['Gen']) else '',
                'Remarque': row['Remarque'] if pd.notna(row['Remarque']) else ''
            }
            
            # Add dual-co information if column exists
            if dual_co_column:
                category_info['DualCo'] = row[dual_co_column] if pd.notna(row[dual_co_column]) else ''
            else:
                category_info['DualCo'] = ''
            
            nename_categories[nename] = category_info
        
        print(f"✅ Loaded categories for {len(nename_categories)} NeNames")
        
        return nename_categories
        
    except Exception as e:
        print(f"❌ Error loading NeName categories: {e}")
        import traceback
        traceback.print_exc()
        return {}
    


def should_feature_be_active(activation_rule, node_type, actual_feature_state, feature_supported, site_gen, site_type, site_cell, site_dual_co):
    """
    Determine if feature should be active based on activation rule, node type, Gen compatibility, Type, Cell and DualCo configuration
    Returns: True if should be active, False if should be inactive, None if rule doesn't apply
    """
    if not activation_rule or not node_type:
        return None
    
    activation_rule = str(activation_rule).lower().strip()
    feature_supported_str = str(feature_supported).lower().strip() if feature_supported else ""
    site_gen_str = str(site_gen).lower().strip() if site_gen else ""
    site_type_str = str(site_type).lower().strip() if site_type else ""
    site_cell_str = str(site_cell).lower().strip() if site_cell else ""
    site_dual_co_str = str(site_dual_co).lower().strip() if site_dual_co else ""
    
    # PRIORITY 1: Check Gen compatibility first (overrides everything)
    if feature_supported_str and site_gen_str:
        # Case: Feature only supports Gen3+ (includes Gen3, Gen4, etc.)
        if ("gen3+" in feature_supported_str or "gen3+=" in feature_supported_str) and "gen2" in site_gen_str:
            return False
        # Case: Feature only supports Gen3 (exactly Gen3, not Gen4)
        elif "gen3" in feature_supported_str and "gen4" in site_gen_str and "gen3+" not in feature_supported_str and "gen3+=" not in feature_supported_str:
            return False
        # Case: Feature only supports Gen4 and site is Gen2/Gen3
        elif "gen4" in feature_supported_str and ("gen2" in site_gen_str or "gen3" in site_gen_str):
            return False
        # Case: Feature supports Gen2 only and site is Gen3/Gen4
        elif "gen2" in feature_supported_str and ("gen3" in site_gen_str or "gen4" in site_gen_str):
            return False
    
    # Split rules by period to handle multiple commands
    rules = [rule.strip() for rule in activation_rule.split('.') if rule.strip()]
    
    node_specific_decisions = []
    general_decisions = []
    
    # Process each rule separately
    for rule in rules:
        # Handle dual-co specific rules - FIXED LOGIC
        if "à la dual-co" in rule or "éligibles à la dual-co" in rule:
            # Check if site has dual-co (non-empty value means it has dual-co)
            if node_type in ['E', 'X'] and site_dual_co_str and site_dual_co_str != "":
                node_specific_decisions.append(True)
            elif node_type in ['E', 'X']:
                node_specific_decisions.append(False)
                
        # Handle complex combined rules first (most specific first)
        
        # Rule: "généralisé en CRZ sur site G + X s'il y TDD"
        elif "généralisé en crz" in rule and "s'il y tdd" in rule and ("site g + x" in rule or "sites g et x" in rule):
            if node_type in ['G', 'X'] and "crz" in site_type_str and ("tdd" in site_cell_str or "tdd+fdd" in site_cell_str or "tdd + fdd" in site_cell_str):
                node_specific_decisions.append(True)
            elif node_type in ['G', 'X']:
                node_specific_decisions.append(False)
                
        # Rule: "A activer sur site G en CRZ"
        elif "a activer sur site g en crz" in rule:
            if node_type == 'G' and "crz" in site_type_str:
                node_specific_decisions.append(True)
            elif node_type == 'G':
                node_specific_decisions.append(False)
                
        # Rule: "A activer sur site G + X s'il ya TDD"
        elif "a activer sur site g + x s'il ya tdd" in rule or "a activer sur site g + x s'il y a tdd" in rule:
            if node_type in ['G', 'X'] and ("tdd" in site_cell_str or "tdd+fdd" in site_cell_str or "tdd + fdd" in site_cell_str):
                node_specific_decisions.append(True)
            elif node_type in ['G', 'X']:
                node_specific_decisions.append(False)
                
        # Rule: "A activer sur sites E et X en ZTD"
        elif "a activer sur sites e et x en ztd" in rule:
            if node_type in ['E', 'X'] and "ztd" in site_type_str:
                node_specific_decisions.append(True)
            elif node_type in ['E', 'X']:
                node_specific_decisions.append(False)
                
        # Rule: "A activer sur sites G et X en CRZ"
        elif "a activer sur sites g et x en crz" in rule:
            if node_type in ['G', 'X'] and "crz" in site_type_str:
                node_specific_decisions.append(True)
            elif node_type in ['G', 'X']:
                node_specific_decisions.append(False)
                
        # Rule: Type-specific rules (CRZ)
        elif "en crz" in rule and not any(term in rule for term in ["site g", "site x", "site e", "sites g et x", "sites e et x"]):
            # General CRZ rule (applies to all nodes in CRZ)
            if "crz" in site_type_str:
                node_specific_decisions.append(True)
            else:
                node_specific_decisions.append(False)
                
        # Rule: Type-specific rules (ZTD)
        elif "en ztd" in rule and not any(term in rule for term in ["site g", "site x", "site e", "sites g et x", "sites e et x"]):
            # General ZTD rule (applies to all nodes in ZTD)
            if "ztd" in site_type_str:
                node_specific_decisions.append(True)
            else:
                node_specific_decisions.append(False)
        
        # Rule: TDD-specific rules
        elif ("s'il ya tdd" in rule or "s'il y a tdd" in rule) and not any(term in rule for term in ["site g", "site x", "site e", "sites g et x", "sites e et x"]):
            # General TDD rule (applies to all nodes with TDD)
            if "tdd" in site_cell_str or "tdd+fdd" in site_cell_str or "tdd + fdd" in site_cell_str:
                node_specific_decisions.append(True)
            else:
                node_specific_decisions.append(False)
        
        # Node-specific activation rules
        elif "a activer sur site x" in rule and node_type == 'X':
            node_specific_decisions.append(True)
        elif "a activer sur site g" in rule and node_type == 'G':
            node_specific_decisions.append(True)
        elif "a activer sur site e" in rule and node_type == 'E':
            node_specific_decisions.append(True)
        elif "a activer sur sites g et x" in rule and node_type in ['G', 'X']:
            node_specific_decisions.append(True)
        elif "a activer sur sites e et x" in rule and node_type in ['E', 'X']:
            node_specific_decisions.append(True)
            
        # Node-specific deactivation rules
        elif "ne pas activer sur site x" in rule and node_type == 'X':
            node_specific_decisions.append(False)
        elif "ne pas activer sur site g" in rule and node_type == 'G':
            node_specific_decisions.append(False)
        elif "ne pas activer sur site e" in rule and node_type == 'E':
            node_specific_decisions.append(False)
            
        # General activation rules (apply to all nodes)
        elif any(activate_rule in rule for activate_rule in [
            "a activer sur les bb configurées en mixed mode", 
            "a activer sur les bb configurées pour supporter mode ess",
            "a activer pour tests",
            "a activer au cas par cas",
            "généralisé en crz",  # Only true general cases without specific nodes
            "a installer",
            "a activer sur sites g et x à partir de"
        ]):
            # Only apply if it's a truly general rule (not combined with other conditions)
            if not any(specific in rule for specific in ["site x", "site g", "site e", "sites g et x", "sites e et x", "s'il ya tdd", "s'il y a tdd", "en crz", "en ztd", "dual-co"]):
                general_decisions.append(True)
                
        # General deactivation rules (apply to all nodes)
        elif any(deactivate_rule in rule for deactivate_rule in [
            "a désactiver sur sites e et x",
            "ne pas activer",
            "ne pas activer par défaut"
        ]):
            # Only apply if it's a general rule (not node-specific)
            if not any(specific in rule for specific in ["site x", "site g", "site e"]):
                general_decisions.append(False)
    
    # Priority: node-specific decisions override general decisions 
    if node_specific_decisions:
        # Return the last node-specific decision (most recent one)
        return node_specific_decisions[-1]
    elif general_decisions:
        # Return the last general decision
        return general_decisions[-1]
    
    # Special case for n/a
    if "n/a" in activation_rule:
        return None
    
    return None


def validate_feature_state(activation_rule, node_type, actual_feature_state, feature_supported, site_gen, site_type, site_cell, site_dual_co):
    """
    Validate if the actual featureState matches what it should be based on activation rule, Gen compatibility, Type, Cell and DualCo
    Returns: "CORRECT", "INCORRECT", or "UNKNOWN"
    """
    expected_active = should_feature_be_active(activation_rule, node_type, actual_feature_state, feature_supported, site_gen, site_type, site_cell, site_dual_co)
    
    if expected_active is None:
        return "UNKNOWN"
    
    # featureState: 1 = active, 0 = inactive
    if expected_active and actual_feature_state == 1:
        return "CORRECT"
    elif not expected_active and actual_feature_state == 0:
        return "CORRECT"
    else:
        return "INCORRECT"

# Get NeName category file path
print("📁 Please provide the NeName category Excel file (Type_Site_ENM.xlsx)")
category_file = input("Enter path to NeName category Excel file: ").strip()
category_path = try_open_excel(category_file)

if category_path is None:
    print(f"❌ NeName category file not found: {category_file}")
    raise SystemExit

# Load NeName categories
nename_categories = load_nename_categories(category_path)
if not nename_categories:
    print("❌ Failed to load NeName categories. Please check the file format.")
    raise SystemExit

# Get parameter file path
print("\n📁 Please provide the parameter Excel file")
param_file = input("Enter path to parameter Excel file (updated file with Features + Licenses sheet): ").strip()
param_path = try_open_excel(param_file)

if param_path is None:
    print(f"❌ Parameter file not found: {param_file}")
    raise SystemExit

# Get data file path
data_file = input("Enter path to data Excel file (with featureState data): ").strip()
data_path = try_open_excel(data_file)

if data_path is None:
    print(f"❌ Data file not found: {data_file}")
    raise SystemExit

print("🔍 Validating Excel files...")
if not is_valid_excel_file(param_path):
    print(f"❌ Parameter file is not a valid Excel file: {param_path}")
    raise SystemExit

if not is_valid_excel_file(data_path):
    print(f"❌ Data file is not a valid Excel file: {data_path}")
    raise SystemExit

print(f"📁 Parameter file: {param_path}")
print(f"📁 Data file: {data_path}")
print(f"📁 NeName category file: {category_path}")

# Load parameter workbook
try:
    print("📖 Loading parameter workbook...")
    param_wb = load_workbook(param_path)
except Exception as e:
    print(f"❌ Error loading parameter workbook: {e}")
    raise SystemExit

# Check if Features + Licenses sheet exists
if "Features + Licenses" not in param_wb.sheetnames:
    print(f"❌ Sheet 'Features + Licenses' not found in parameter file")
    print(f"Available sheets: {param_wb.sheetnames}")
    raise SystemExit

# Read Features + Licenses sheet
features_ws = param_wb["Features + Licenses"]
features_header = [cell.value for cell in features_ws[1]]

print(f"📋 Columns in Features + Licenses sheet: {features_header}")

# Map column names to indices
features_col_idx_map = {name: idx + 1 for idx, name in enumerate(features_header) if name is not None}

# Check required columns in Features + Licenses file
required_features_columns = ["Feature name", "Bytel nodes", "BB / DU supported", "FeatureState", "A activer ou pas pour Bytel"]

missing_columns = []
for col in required_features_columns:
    if col not in features_col_idx_map:
        missing_columns.append(col)

if missing_columns:
    print(f"❌ Missing required columns in Features + Licenses sheet: {missing_columns}")
    raise SystemExit

# Get features data
features_data = []
feature_name_col = features_col_idx_map["Feature name"]
feature_state_col = features_col_idx_map["FeatureState"]
feature_supported_col = features_col_idx_map["BB / DU supported"]
activate_col = features_col_idx_map["A activer ou pas pour Bytel"]
nodes_bytel = features_col_idx_map["Bytel nodes"]

print(f"\n🔍 Collecting features from 'Features + Licenses' sheet...")

for row in range(2, features_ws.max_row + 1):
    feature_name = features_ws.cell(row=row, column=feature_name_col).value
    feature_state = features_ws.cell(row=row, column=feature_state_col).value
    activate = features_ws.cell(row=row, column=activate_col).value
    nodes = features_ws.cell(row=row, column=nodes_bytel).value
    supported = features_ws.cell(row=row, column=feature_supported_col).value

    if feature_name is not None and str(feature_name).strip() != "":
        feature_clean = str(feature_name).strip()

        features_data.append({
            "Feature name": feature_clean,
            "FeatureState": feature_state,
            "A activer ou pas pour Bytel": activate,
            "Bytel nodes": nodes,
            "BB / DU supported" : supported
        })

print(f"✅ Collected {len(features_data)} features from Features + Licenses sheet")

# Load data workbook
try:
    print(f"\n📖 Loading data workbook...")
    data_df = pd.read_excel(data_path, engine="openpyxl")
except Exception as e:
    print(f"❌ Error loading data file: {e}")
    raise SystemExit

print(f"📊 Data file shape: {data_df.shape}")
print(f"📋 Data file columns: {list(data_df.columns)}")

# Check if data file has required columns
required_data_columns = ["featureStateId", "NeName", "featureState", "serviceState"]
missing_data_columns = []
for col in required_data_columns:
    if col not in data_df.columns:
        missing_data_columns.append(col)

if missing_data_columns:
    print(f"❌ Missing required columns in data file: {missing_data_columns}")
    raise SystemExit

# Create output data
output_data = []
incorrect_data = []

print(f"\n📝 Creating license validation table...")

# Define color fills
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
YELLOW_FILL2 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for not found

# Process each feature from the parameter file
for feature_info in features_data:
    feature_name = feature_info["Feature name"]
    feature_state_id = feature_info["FeatureState"]
    activate_for_bytel = feature_info["A activer ou pas pour Bytel"]
    nodes_for_bytel = feature_info["Bytel nodes"]
    feature_supported = feature_info["BB / DU supported"]

    matching_rows = data_df[data_df["featureStateId"] == feature_state_id]

    if len(matching_rows) > 0:
        for _, row in matching_rows.iterrows():
            site_name = row["NeName"]
            actual_feature_state = row["featureState"]
            service_state = row["serviceState"]
            
            node_type = get_node_type(site_name)
            category_info = nename_categories.get(str(site_name).strip(), {})
            site_gen = category_info.get('Gen', '')
            site_type = category_info.get('Type', '')
            site_cell = category_info.get('Cell', '')
            site_dual_co = category_info.get('DualCo', '')  # Get dual-co information
            
            # Pass all information to validation including dual-co
            validation_status = validate_feature_state(
                activate_for_bytel, 
                node_type, 
                actual_feature_state, 
                feature_supported, 
                site_gen,
                site_type,
                site_cell,
                site_dual_co  # Pass dual-co information
            )
            
            output_row = {
                "Feature name": feature_name,
                "FeatureState": feature_state_id,
                "Bytel nodes": nodes_for_bytel,
                "BB / DU supported": feature_supported,
                "A activer ou pas pour Bytel": activate_for_bytel,
                "NeName": site_name,
                "featureState": actual_feature_state,
                "serviceState": service_state,
                "NodeType": node_type,
                "Type": site_type,
                "Operateur": category_info.get('Operateur', ''),
                "Cell": site_cell,
                "Gen": site_gen,
                "DualCo": site_dual_co,  # Add dual-co column to output
                "Remarque": category_info.get('Remarque', ''),
                "Validation": validation_status
            }
            
            output_data.append(output_row)
            
            if validation_status == "INCORRECT":
                incorrect_data.append(output_row)

# Create output DataFrame
output_df = pd.DataFrame(output_data)
incorrect_df = pd.DataFrame(incorrect_data)

print(f"✅ Created output table with {len(output_df)} rows")
print(f"❌ Found {len(incorrect_df)} incorrect feature states")

# Create output file name
output_filename = "License_Validation_Report.xlsx"

# Save to new Excel file
try:
    print(f"\n💾 Saving output to: {output_filename}")

    # Create a new workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create main validation sheet
    ws_main = wb.create_sheet("License_Validation")

    # Write headers for main sheet
    main_headers = ["Feature name", "FeatureState", "Bytel nodes", "BB / DU supported", "A activer ou pas pour Bytel",
                    "NeName", "featureState", "serviceState", "NodeType", "Type", "Operateur", 
                            "Cell", "Gen", "DualCo", "Remarque", "Validation"]

    for col_idx, header in enumerate(main_headers, 1):
        ws_main.cell(row=1, column=col_idx, value=header)

    # Write data and apply formatting
    current_feature = None
    merge_start_row = 2

    for row_idx, (_, row_data) in enumerate(output_df.iterrows(), 2):
        # Write row data
        for col_idx, header in enumerate(main_headers, 1):
            ws_main.cell(row=row_idx, column=col_idx, value=row_data[header])

        # Apply color coding to columns starting from NeName (column 5) for incorrect entries
        validation_status = row_data["Validation"]
        if validation_status == "INCORRECT":
            # Highlight columns from NeName to Validation (columns 5 to 8) in red
            for col in range(6, len(main_headers) + 1):  # Columns 5 to 8
                ws_main.cell(row=row_idx, column=col).fill = YELLOW_FILL
        elif validation_status == "NOT FOUND":
            # Highlight columns from NeName to Validation (columns 5 to 8) in yellow for not found
            for col in range(5, len(main_headers) + 1):  # Columns 5 to 8
                ws_main.cell(row=row_idx, column=col).fill = YELLOW_FILL

        # Check if this is a new feature group
        if row_data["Feature name"] != current_feature:
            # If we were tracking a previous feature, merge its cells
            if current_feature is not None and merge_start_row < row_idx - 1:
                for col in range(1, 5):  # Merge columns A to C (Feature name, FeatureState, A activer ou pas)
                    ws_main.merge_cells(start_row=merge_start_row, start_column=col,
                                        end_row=row_idx - 1, end_column=col)

            # Start tracking new feature
            current_feature = row_data["Feature name"]
            merge_start_row = row_idx

    # Merge the last feature group
    if current_feature is not None and merge_start_row < len(output_df) + 1:
        for col in range(1, 4):  # Merge columns A to C
            ws_main.merge_cells(start_row=merge_start_row, start_column=col,
                                end_row=len(output_df) + 1, end_column=col)

    # Create incorrect entries sheet
    ws_incorrect = wb.create_sheet("Incorrect_Entries")
    
    if len(incorrect_df) > 0:
        # Write headers for incorrect sheet
        incorrect_headers = ["Feature name", "FeatureState", "Bytel nodes", "BB / DU supported", "A activer ou pas pour Bytel",
                            "NeName", "featureState", "serviceState", "NodeType", "Type", "Operateur", 
                            "Cell", "Gen", "DualCo", "Remarque", "Validation"]
        
        for col_idx, header in enumerate(incorrect_headers, 1):
            ws_incorrect.cell(row=1, column=col_idx, value=header)
        
        # Write incorrect data
        for row_idx, (_, row_data) in enumerate(incorrect_df.iterrows(), 2):
            for col_idx, header in enumerate(incorrect_headers, 1):
                ws_incorrect.cell(row=row_idx, column=col_idx, value=row_data[header])
            
            # Highlight columns from NeName to Validation (columns 5 to 8) in red for incorrect entries
            for col in range(6, len(incorrect_headers) + 1):  # Columns 5 to 8
                ws_incorrect.cell(row=row_idx, column=col).fill = YELLOW_FILL
    else:
        ws_incorrect.cell(row=1, column=1, value="No incorrect entries found!")
    
    # Create summary sheet
    ws_summary = wb.create_sheet("Summary")

    # Calculate summary statistics
    total_features = len(features_data)
    total_entries = len(output_df)

    features_found = len(output_df[output_df["NeName"] != "NOT FOUND"])
    features_not_found = len(output_df[output_df["NeName"] == "NOT FOUND"])

    correct_entries = len(output_df[output_df["Validation"] == "CORRECT"])
    incorrect_entries = len(output_df[output_df["Validation"] == "INCORRECT"])
    unknown_entries = len(output_df[output_df["Validation"] == "UNKNOWN"])
    not_found_entries = len(output_df[output_df["Validation"] == "NOT FOUND"])
    
    unique_sites = output_df[output_df["NeName"] != "NOT FOUND"]["NeName"].nunique()

    # Write summary
    ws_summary.cell(row=1, column=1, value="License Validation Summary").font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:B1')

    summary_data = [
        ("Total Features in Parameter File", total_features),
        ("Total Validation Entries", total_entries),
        ("Unique Sites Found", unique_sites),
        ("Features Found in Data File", features_found),
        ("Features Not Found in Data File", features_not_found),
        ("✅ CORRECT Feature States", correct_entries),
        ("❌ INCORRECT Feature States", incorrect_entries),
        ("⚪ UNKNOWN Feature States", unknown_entries),
        ("🔍 NOT FOUND Features", not_found_entries),
    ]

    for i, (label, value) in enumerate(summary_data, 3):
        ws_summary.cell(row=i, column=1, value=label)
        ws_summary.cell(row=i, column=2, value=value)

    # Add features not found section
    start_row = len(summary_data) + 5
    ws_summary.cell(row=start_row, column=1, value="Features Not Found in Data File").font = Font(bold=True)
    ws_summary.merge_cells(f'A{start_row}:B{start_row}')

    not_found_features = []
    for feature_info in features_data:
        feature_state_id = feature_info["FeatureState"]
        if feature_state_id not in data_df["featureStateId"].values:
            not_found_features.append(f"{feature_info['Feature name']} (FeatureState: {feature_state_id})")

    if not_found_features:
        for i, feature in enumerate(not_found_features, start_row + 1):
            ws_summary.cell(row=i, column=1, value=feature)
    else:
        ws_summary.cell(row=start_row + 1, column=1, value="All features were found in data file")

    # Auto-adjust column widths for all sheets
    sheets_to_adjust = [ws_main, ws_incorrect, ws_summary]

    for ws in sheets_to_adjust:
        for column_cells in ws.columns:
            # Skip if it's a MergedCell (which doesn't have column_letter attribute)
            if isinstance(column_cells[0], MergedCell):
                continue

            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    # Skip MergedCell objects
                    if isinstance(cell, MergedCell):
                        continue
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Add borders and formatting to all sheets
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for ws in sheets_to_adjust:
        if ws == ws_main:
            max_row = len(output_df) + 1
            max_col = len(main_headers)
        elif ws == ws_incorrect:
            max_row = len(incorrect_df) + 1 if len(incorrect_df) > 0 else 1
            max_col = len(incorrect_headers) if len(incorrect_df) > 0 else 1
        elif ws == ws_summary:
            max_row = start_row + len(not_found_features) if not_found_features else start_row + 1
            max_col = 2
        else:
            continue

        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                # Only apply border to non-merged cells
                if not isinstance(cell, MergedCell):
                    cell.border = thin_border

        # Style header row (skip merged cells)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            if not isinstance(cell, MergedCell):
                cell.fill = header_fill
                cell.font = Font(bold=True)

    # Add legend to main sheet
    legend_row = len(output_df) + 3
    ws_main.cell(row=legend_row, column=1, value="LEGEND:").font = Font(bold=True)
    ws_main.cell(row=legend_row + 1, column=1, value="❌ RED highlighting").fill = YELLOW_FILL
    ws_main.cell(row=legend_row + 1, column=2, value="= Incorrect feature state (from NeName to Validation)")
    ws_main.cell(row=legend_row + 2, column=1, value="🔍 YELLOW highlighting").fill = YELLOW_FILL2
    ws_main.cell(row=legend_row + 2, column=2, value="= Feature not found in data file (from NeName to Validation)")

    wb.save(output_filename)

    print(f"🎉 SUCCESS!")
    print(f"📁 Output file created: {output_filename}")
    print(f"📊 Sheets created:")
    print(f"   - License_Validation: Main validation results (NeName to Validation highlighted for issues)")
    print(f"   - Incorrect_Entries: List of all incorrect feature states")
    print(f"   - Summary: Overall validation summary")
    print(f"📊 Final Summary:")
    print(f"   📋 Total Features: {total_features}")
    print(f"   🏢 Unique Sites: {unique_sites}")
    print(f"   ⚪ UNKNOWN Feature States: {unknown_entries}")
    print(f"   🔍 NOT FOUND Features: {not_found_entries}")

    # Print incorrect features if any


except Exception as e:
    print(f"❌ Error saving output file: {e}")
    import traceback
    traceback.print_exc()
