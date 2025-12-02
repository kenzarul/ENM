import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

choice = input("Select target sheet (NRCellCU / NRCellDU): ").strip().upper()
# ---------- CONFIG ----------
SHEET_NAME = "LTE - NR parameters" # sheet to read (you said this name)
if choice == "NRCELLCU":
    SHEET_NAME2 = "NRCellCU" # sheet to read (you said this name)
elif choice == "NRCELLDU":
    SHEET_NAME2 = "NRCellDU" # sheet to read (you said this name)
PARAM_COL_NAME = "Parameter" # exact header name for parameter column
COLUMNS_TO_COMPARE = ["lock / unlock", "Valeur par défaut RBS", "Valeur Bytel TDD MidBand", "Valeur Bytel FDD ESS 15MHz", 
                      "Valeur Bytel TDD+FDD co-node\nAppliquer la valeur commune si valeur TDD et FDD sont même, sinon appliquer la valeur spécifiée dans cette colonne.", 
                      "Valeur Bytel TDD HigBand", "Commentaire", "Delta 25.Q1 E//", "Comment 25.Q1 E//", "Delta 25.Q2 E//", "Comment 25.Q2 E//"] # add any other exact column names
# -----------------------------

def try_open_excel(path):
    # try as given, otherwise try common extensions
    candidates = [path]
    if not os.path.splitext(path)[1]:
        candidates += [path + ext for ext in (".xlsx", ".xls", ".xlsm")]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None

def find_new_parameters_from_lab(old_file_path, lab_param_file_path, param_column_name):
    """
    Compare LAB parameters with OLD file parameters and return parameters that exist in LAB but not in OLD file.
    """
    try:
        # Read OLD file parameters
        old_df = pd.read_excel(old_file_path, sheet_name=SHEET_NAME2, engine="openpyxl")
        old_df[param_column_name] = old_df[param_column_name].astype(str).str.strip()
        old_params_set = set(old_df[param_column_name].str.lower())
        
        # Read LAB parameters file
        with open(lab_param_file_path, "r", encoding="utf-8") as f:
            lab_params = [line.strip() for line in f if line.strip()]
        
        # Find parameters in LAB but not in OLD
        missing_in_old = []
        for lab_param in lab_params:
            lab_param_clean = lab_param.strip()
            lab_param_lower = lab_param_clean.lower()
            
            if lab_param_lower not in old_params_set:
                missing_in_old.append(lab_param_clean)
        
        return missing_in_old
        
    except Exception as e:
        print(f"❌ Error comparing with LAB parameters: {e}")
        return []

old_file = input("Enter path to Excel file VRTO : ").strip()
new_file = input("Enter path to Excel file LTE + NR : ").strip()
param_list_file = input("Enter path to parameter list TXT file from Excel file: ").strip()

old_path = try_open_excel(old_file)
new_path = try_open_excel(new_file)

if old_path is None:
    print(f"❌ OLD file not found. Tried: {old_file} and common extensions (.xlsx/.xls/.xlsm).")
    raise SystemExit
if new_path is None:
    print(f"❌ UPDATED file not found. Tried: {new_file} and common extensions (.xlsx/.xls/.xlsm).")
    raise SystemExit

# Read sheet by name
try:
    print("Sheets in Old:", pd.ExcelFile(old_path).sheet_names)
    old_df = pd.read_excel(old_path, sheet_name=SHEET_NAME2, engine="openpyxl")
except Exception as e:
    print(f"❌ Could not read sheet '{SHEET_NAME2}' from OLD file: {old_path}")
    print("Error:", e)
    print("Available sheets (OLD):", pd.ExcelFile(old_path, engine="openpyxl").sheet_names)
    raise SystemExit

try:
    print("Sheets in New:", pd.ExcelFile(new_path).sheet_names)
    new_df = pd.read_excel(new_path, sheet_name=SHEET_NAME, engine="openpyxl")
except Exception as e:
    print(f"❌ Could not read sheet '{SHEET_NAME}' from UPDATED file: {new_path}")
    print("Error:", e)
    print("Available sheets (UPDATED):", pd.ExcelFile(new_path, engine="openpyxl").sheet_names)
    raise SystemExit

# Read parameter list
if not os.path.exists(param_list_file):
    print(f"❌ Parameter list file not found: {param_list_file}")
    raise SystemExit

with open(param_list_file, "r", encoding="utf-8") as f:
    parameters = [line.strip() for line in f if line.strip()]

# Ensure parameter column exists
if PARAM_COL_NAME not in old_df.columns or PARAM_COL_NAME not in new_df.columns:
    print(f"❌ Column '{PARAM_COL_NAME}' not found in one of the sheets.")
    print("OLD columns:", list(old_df.columns))
    print("NEW columns:", list(new_df.columns))
    raise SystemExit

# Make lookups by trimming parameter strings
old_df[PARAM_COL_NAME] = old_df[PARAM_COL_NAME].astype(str).str.strip()
new_df[PARAM_COL_NAME] = new_df[PARAM_COL_NAME].astype(str).str.strip()

old_lookup = old_df.set_index(PARAM_COL_NAME)
new_lookup = new_df.set_index(PARAM_COL_NAME)

differences = []
missing = []
for p in parameters:
    p = p.strip()
    in_old = p in old_lookup.index
    in_new = p in new_lookup.index
    if not in_old or not in_new:
        missing.append({"Parameter": p, "In_old": in_old, "In_new": in_new})
        continue
    for col in COLUMNS_TO_COMPARE:
        if col not in old_lookup.columns or col not in new_lookup.columns:
            print(f"❌ Compare column '{col}' not found in one of the sheets.")
            print("OLD columns:", list(old_df.columns))
            print("NEW columns:", list(new_df.columns))
            raise SystemExit
        old_val = old_lookup.at[p, col]
        if isinstance(old_val, pd.Series):
            old_val = old_val.iloc[0]
        new_val = new_lookup.at[p, col]
        if isinstance(new_val, pd.Series):
            new_val = new_val.iloc[0]
        # normalize NaN vs None
        if pd.isna(old_val) and pd.isna(new_val):
            continue
        if (pd.isna(old_val) and not pd.isna(new_val)) or (pd.isna(new_val) and not pd.isna(old_val)) or (old_val != new_val):
            differences.append({
                "Parameter": p,
                "Column": col,
                "Old Value": old_val,
                "New Value": new_val
            })

# Find parameters present in new but not old (new parameters)
new_only = []
for p in parameters:
    p_lower = p.strip().lower()
    old_index_set = set(old_lookup.index.str.strip().str.lower())
    new_index_set = set(new_lookup.index.str.strip().str.lower())
    
    if p_lower in new_index_set and p_lower not in old_index_set:
        new_only.append({"Parameter": p})

# Print summary
print("\n=== SUMMARY ===")
print(f"Parameters checked from list: {len(parameters)}")
print(f"Differences found: {len(differences)}")
print(f"Parameters missing in one file: {len(missing)}")
print(f"New parameters (present in UPDATED but not in OLD): {len(new_only)}\n")

if differences:
    print("== Differences ==")
    for d in differences:
        print(f"[DIFF] {d['Parameter']} | {d['Column']}: OLD='{d['Old Value']}' -> NEW='{d['New Value']}'")

if missing:
    print("\n== Missing Parameters ==")
    for m in missing:
        print(f"{m['Parameter']} in_old={m['In_old']} in_new={m['In_new']}")

if new_only:
    print("\nNew parameters found in UPDATED file:")
    for n in new_only:
        print(n['Parameter'])
        
for col in COLUMNS_TO_COMPARE:
    if col in old_df.columns:
        old_df[col] = old_df[col].astype(object)

# Load the workbook and target sheet
wb = load_workbook(old_path)
ws = wb[SHEET_NAME2]

purple_fill = PatternFill(start_color="CBC3E3", end_color="CBC3E3", fill_type="solid")

# Find the header row (assume first row)
header = [cell.value for cell in ws[1]]

# Map column names to their index in Excel
col_idx_map = {name: idx+1 for idx, name in enumerate(header)}

# Ensure the columns we want to update exist
for col in COLUMNS_TO_COMPARE:
    if col not in col_idx_map:
        raise ValueError(f"Column '{col}' not found in sheet '{SHEET_NAME2}'")

if PARAM_COL_NAME not in col_idx_map:
    raise ValueError(f"Parameter column '{PARAM_COL_NAME}' not found in sheet '{SHEET_NAME2}'")

param_col_idx = col_idx_map[PARAM_COL_NAME]

# Update only the cells that differ
for row in range(2, ws.max_row + 1):
    param_value = ws.cell(row=row, column=param_col_idx).value
    if param_value is None:
        continue
    param_value_str = str(param_value).strip()
    
    # Check if this parameter has a difference
    for diff in differences:
        if diff["Parameter"].strip() == param_value_str:
            col_idx = col_idx_map[diff["Column"]]
            cell = ws.cell(row=row, column=col_idx)
            existing_comment = cell.comment
            cell.value = diff["New Value"]
            cell.fill = purple_fill
            if existing_comment:
                cell.comment = existing_comment
            # Only overwrite the differing cell
            break  # done with this row

# Save workbook in-place
wb.save(old_path)
print(f"\n✅ OLD Excel updated in place — only differing cells in {COLUMNS_TO_COMPARE} updated.")

# Ask user if they want to check for new parameters from LAB
print("\n" + "="*60)
check_lab = input("\nDo you want to check for new parameters from LAB? (yes/no): ").strip().lower()

if check_lab in ['yes', 'y', 'oui', 'o']:
    lab_param_file = input("Enter path to LAB parameter list TXT file: ").strip()
    
    if os.path.exists(lab_param_file):
        print("\n🔍 Checking for parameters in LAB that are not in OLD file...")
        
        # Find parameters that exist in LAB but not in OLD file
        missing_in_old = find_new_parameters_from_lab(old_path, lab_param_file, PARAM_COL_NAME)
        
        if missing_in_old:
            print(f"\n📋 Found {len(missing_in_old)} parameters in LAB that are NOT in OLD file:")
            print("-" * 80)
            for i, param in enumerate(missing_in_old, 1):
                print(f"{i:3}. {param}")
            print("-" * 80)

        else:
            print("\n✅ All LAB parameters are already present in the OLD file.")
    else:
        print(f"❌ LAB parameter file not found: {lab_param_file}")

print("\n🎯 Script execution completed!")
